from datetime import datetime, timedelta
import os
import json
import zipfile
import urllib.parse
import xml.etree.ElementTree as ET
import mimetypes

from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Max, Count, OuterRef, Subquery, Exists
from django.db import transaction
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from django.urls import reverse
from django.conf import settings
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.files.storage import FileSystemStorage
from django.utils.text import get_valid_filename
from django.http import Http404, HttpResponse, HttpResponseForbidden
from django.http import FileResponse
from django.http import HttpResponsePermanentRedirect
from django.views.decorators.clickjacking import xframe_options_sameorigin

from .models import (
    BusinessTenant,
    ContractorDocument,
    ContractorProfile,
    EmployeeProfile,
    JobTitle,
    TrainerProfile,
)
from .forms import (
    BusinessEmployeeCreateForm,
    CourseAssignmentRuleForm,
    CourseForm,
    JobTitleForm,
    RegisterForm,
    SOPChecklistAssignmentRuleForm,
    SOPChecklistForm,
    ExamTemplateForm,
    ManualQuestionFormSet,
    ExcelQuestionsUploadForm,

    # ✅✅✅ الحل: فورم السؤال + inline formset للخيارات (يحفظ في ExamOption model)
    ExamQuestionForm,
    ExamOptionInlineFormSet,

    # Super Admin
    ProgramForm,

    # Training Coordinator
    TrainingCoordinatorRegisterContractorForm,
)

from training.models import (
    Course,
    CourseAssignment,
    CourseAssignmentRule,
    Program,
    EnrollmentRequest,
    ExamSession,
    ExamTemplate,
    ExamQuestion,
    ExamOption,
    ExamAttempt,
    ProgramExamPartConfig,
    ExternalPartAssessment,
    SOPChecklist,
    SOPChecklistAssignmentRule,
    SOPChecklistCompletion,
    SOPChecklistItem,
    SOPChecklistItemCompletion,
)

User = get_user_model()

import io
import uuid
from django.http import JsonResponse
from django.core.files.base import ContentFile

from certification.models import Certificate, ScormCertificate  # ✅ تأكد من اسم التطبيق


def _parse_ymd_date(value: str | None):
    """Parse YYYY-MM-DD into date, return None if invalid."""
    if not value:
        return None
    value = (value or '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def _date_range_to_aware_datetimes(start_date, end_date):
    """Convert date range to inclusive aware datetime range [start, end]."""
    if not start_date and not end_date:
        return None, None

    tz = timezone.get_current_timezone()
    start_dt = None
    end_dt = None

    if start_date:
        start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()), tz)
    if end_date:
        end_dt = timezone.make_aware(datetime.combine(end_date, datetime.max.time().replace(microsecond=0)), tz)

    return start_dt, end_dt


# Ensure common SCORM asset types are served with correct MIME types (Windows may miss some mappings)
mimetypes.add_type('text/css', '.css')
mimetypes.add_type('application/javascript', '.js')
mimetypes.add_type('application/json', '.json')
mimetypes.add_type('image/svg+xml', '.svg')
mimetypes.add_type('audio/mpeg', '.mp3')


# =========================
# Helpers
# =========================

def _is_contractor(user) -> bool:
    # A user must not be both contractor and trainer.
    # If they have trainer role, treat them as trainer (not contractor).
    if not user:
        return False
    if TrainerProfile.objects.filter(user=user).exists():
        return False
    # Training coordinator is a separate role (not a contractor).
    return ContractorProfile.objects.filter(user=user, is_training_coordinator=False).exists()


def _is_trainer(user) -> bool:
    if not user:
        return False
    return TrainerProfile.objects.filter(user=user).exists()


def _is_training_coordinator(user) -> bool:
    if not user:
        return False
    # Coordinator is a contractor with a flag, and must not be a trainer.
    if TrainerProfile.objects.filter(user=user).exists():
        return False
    return ContractorProfile.objects.filter(user=user, is_training_coordinator=True).exists()


def _is_business_owner(user) -> bool:
    if not user:
        return False
    return BusinessTenant.objects.filter(owner=user, is_active=True).exists()


def _is_employee(user) -> bool:
    if not user:
        return False
    return EmployeeProfile.objects.filter(user=user, is_active=True, business__is_active=True).exists()


def _training_coordinator_guard(request) -> bool:
    user = getattr(request, 'user', None)
    return bool(user and user.is_authenticated and _is_training_coordinator(user))


def _business_owner_guard(request) -> bool:
    user = getattr(request, 'user', None)
    return bool(user and user.is_authenticated and (_is_business_owner(user) or _ensure_legacy_business_owner(user)))


def _employee_guard(request) -> bool:
    user = getattr(request, 'user', None)
    return bool(user and user.is_authenticated and _is_employee(user))


def _is_super_admin(user) -> bool:
    return bool(getattr(user, 'is_superuser', False))


def _super_admin_guard(request) -> bool:
    if not _is_super_admin(getattr(request, 'user', None)):
        messages.error(request, 'غير مصرح لك بالدخول')
        return False
    return True


def _get_owned_business(user):
    return BusinessTenant.objects.filter(owner=user, is_active=True).first()


def _ensure_legacy_business_owner(user):
    if not user or _is_super_admin(user) or _is_business_owner(user) or _is_employee(user):
        return _get_owned_business(user)

    legacy_profile = ContractorProfile.objects.filter(user=user).first()
    if legacy_profile is None:
        return None

    business_name = (legacy_profile.company_name or '').strip() or (user.get_full_name() or '').strip() or user.username
    business, _created = BusinessTenant.objects.get_or_create(
        owner=user,
        defaults={
            'name': business_name,
            'industry': 'Food & Beverage',
        },
    )
    return business


def _primary_dashboard_route(user):
    if not user or not user.is_authenticated:
        return 'home'
    if _is_super_admin(user):
        return 'super_admin_dashboard'
    if _is_business_owner(user):
        return 'business_owner_dashboard'
    if _ensure_legacy_business_owner(user):
        return 'business_owner_dashboard'
    if _is_employee(user):
        return 'employee_dashboard'
    return 'home'


def _get_employee_profile(user):
    return (
        EmployeeProfile.objects
        .select_related('business', 'job_title', 'user')
        .filter(user=user, is_active=True, business__is_active=True)
        .first()
    )


def _provision_course_assignments_for_employee(employee_profile, assigned_by=None):
    if not employee_profile or not employee_profile.job_title_id:
        return

    single_job_title_mode = (
        JobTitle.objects
        .filter(business=employee_profile.business)
        .count() == 1
    )
    rules = (
        CourseAssignmentRule.objects
        .filter(
            business=employee_profile.business,
            job_title=employee_profile.job_title,
            course__is_active=True,
        )
        .select_related('course', 'job_title')
    )
    if single_job_title_mode:
        assigned_course_ids = set(rules.values_list('course_id', flat=True))
        implicit_courses = (
            Course.objects
            .filter(business=employee_profile.business, is_active=True)
            .exclude(id__in=assigned_course_ids)
        )
        for course in implicit_courses:
            CourseAssignment.objects.get_or_create(
                business=employee_profile.business,
                course=course,
                employee=employee_profile.user,
                defaults={
                    'assigned_by': assigned_by,
                    'assigned_via_job_title': employee_profile.job_title,
                },
            )
    for rule in rules:
        CourseAssignment.objects.get_or_create(
            business=employee_profile.business,
            course=rule.course,
            employee=employee_profile.user,
            defaults={
                'assigned_by': assigned_by or rule.assigned_by,
                'assigned_via_job_title': rule.job_title,
            },
        )


def _ensure_course_assignments_for_rule(rule):
    employees = (
        EmployeeProfile.objects
        .filter(
            business=rule.business,
            job_title=rule.job_title,
            is_active=True,
            user__is_active=True,
        )
        .select_related('user')
    )
    for employee in employees:
        CourseAssignment.objects.get_or_create(
            business=rule.business,
            course=rule.course,
            employee=employee.user,
            defaults={
                'assigned_by': rule.assigned_by,
                'assigned_via_job_title': rule.job_title,
            },
        )


def _assigned_checklists_queryset(employee_profile):
    if not employee_profile or not employee_profile.job_title_id:
        return SOPChecklist.objects.none()

    queryset = (
        SOPChecklist.objects
        .filter(
            business=employee_profile.business,
            is_active=True,
            assignment_rules__job_title=employee_profile.job_title,
        )
        .prefetch_related('items')
        .distinct()
        .order_by('title', 'id')
    )
    if queryset.exists():
        return queryset

    single_job_title_mode = (
        JobTitle.objects
        .filter(business=employee_profile.business)
        .count() == 1
    )
    if single_job_title_mode:
        return (
            SOPChecklist.objects
            .filter(
                business=employee_profile.business,
                is_active=True,
            )
            .prefetch_related('items')
            .order_by('title', 'id')
        )

    return queryset


def _handle_scorm_upload_post(request, success_redirect_name: str):
    storage = _scorm_storage()
    uploaded = request.FILES.get('scorm_zip')
    if not uploaded:
        messages.error(request, 'يرجى اختيار ملف SCORM بصيغة ZIP')
        return redirect(success_redirect_name)

    if not uploaded.name.lower().endswith('.zip'):
        messages.error(request, 'صيغة الملف غير مدعومة. يرجى رفع ملف ZIP فقط')
        return redirect(success_redirect_name)

    safe_name = get_valid_filename(uploaded.name)
    if not safe_name.lower().endswith('.zip'):
        safe_name = safe_name + '.zip'

    candidate = safe_name
    if storage.exists(candidate):
        base, ext = os.path.splitext(safe_name)
        candidate = f"{base}_{uuid.uuid4().hex[:8]}{ext}"

    storage.save(candidate, uploaded)

    # حاول استخراج الملف وتجهيزه للتشغيل
    try:
        _ensure_extracted_for_zip(candidate)
    except Exception:
        pass

    messages.success(request, 'تم رفع ملف SCORM بنجاح')
    return redirect(success_redirect_name)


def _safe_program_name(program) -> str:
    if not program:
        return '—'

    perceived = getattr(program, 'title', None) or getattr(program, 'name', None)
    return str(perceived) if perceived else str(program)


def _english_text_only(value: object, fallback: str = 'N/A') -> str:
    """
    Keep printable ASCII characters only (English-friendly output for PDF fields).
    """
    raw = '' if value is None else str(value)
    ascii_text = ''.join(ch for ch in raw if 32 <= ord(ch) <= 126)
    ascii_text = ' '.join(ascii_text.split())
    return ascii_text or fallback


def _contractor_english_name(user) -> str:
    # Registration stores English full name in first_name.
    # Requirement: never use username on certificates/pass cards.
    first_name_en = _english_text_only(getattr(user, 'first_name', ''), fallback='')
    if first_name_en:
        return first_name_en

    candidates = [
        getattr(user, 'get_full_name', lambda: '')(),
        f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip(),
    ]
    for c in candidates:
        cleaned = _english_text_only(c, fallback='')
        if cleaned:
            return cleaned
    return 'Contractor'


def _program_english_name(program) -> str:
    if not program:
        return 'Program'
    return _english_text_only(
        getattr(program, 'title_en', None) or getattr(program, 'title', None) or getattr(program, 'name', None),
        fallback='Program',
    )


def _get_exam_session_key(enrollment_id: int) -> str:
    return f"exam_answers_enrollment_{enrollment_id}"


def _get_exam_code_unlock_session_key(enrollment_id: int) -> str:
    return f"exam_code_unlocked_{enrollment_id}"


def _normalize_q_index(value, total: int) -> int:
    try:
        i = int(value)
    except (TypeError, ValueError):
        i = 1

    if i < 1:
        i = 1
    if total and i > total:
        i = total
    return i


MAX_EXAM_ATTEMPTS_PER_PART = 3
PART_ATTEMPT_FIELD_BY_TYPE = {
    ProgramExamPartConfig.PartType.THEORETICAL: 'theoretical_attempts_count',
    ProgramExamPartConfig.PartType.PRACTICAL: 'practical_attempts_count',
    ProgramExamPartConfig.PartType.PROJECT: 'project_attempts_count',
}


def _get_part_attempts_count(enrollment: EnrollmentRequest, part_type: str) -> int:
    field_name = PART_ATTEMPT_FIELD_BY_TYPE.get(part_type)
    if not field_name:
        return 0
    return int(getattr(enrollment, field_name, 0) or 0)


def _increment_part_attempts_count(enrollment: EnrollmentRequest, part_type: str) -> tuple[str | None, int]:
    field_name = PART_ATTEMPT_FIELD_BY_TYPE.get(part_type)
    if not field_name:
        return None, 0
    next_count = _get_part_attempts_count(enrollment, part_type) + 1
    setattr(enrollment, field_name, next_count)
    return field_name, next_count


def _part_passing_threshold(program: Program, part_type: str, default: int = 60) -> int:
    conf = (
        ProgramExamPartConfig.objects
        .filter(program=program, part_type=part_type, is_required=True)
        .order_by('order', 'id')
        .first()
    )
    if conf is not None and getattr(conf, 'passing_grade_percent', None) is not None:
        return int(conf.passing_grade_percent)
    return int(default)


def _has_passed_part(enrollment: EnrollmentRequest, part_type: str) -> bool:
    if part_type == ProgramExamPartConfig.PartType.THEORETICAL:
        return bool(enrollment.has_passed_theoretical())
    if part_type == ProgramExamPartConfig.PartType.PRACTICAL:
        return bool(enrollment.has_passed_practical())
    if part_type == ProgramExamPartConfig.PartType.PROJECT:
        return bool(enrollment.has_passed_project())
    return False


def _is_locked_after_max_part_attempts(enrollment: EnrollmentRequest) -> bool:
    required_parts = set(enrollment.required_exam_parts() or [])
    for part_type in required_parts:
        attempts = _get_part_attempts_count(enrollment, part_type)
        # Backward-compatible: old data used attempts_count for theoretical only.
        if part_type == ProgramExamPartConfig.PartType.THEORETICAL:
            attempts = max(attempts, int(getattr(enrollment, 'attempts_count', 0) or 0))
        if attempts >= MAX_EXAM_ATTEMPTS_PER_PART and not _has_passed_part(enrollment, part_type):
            return True
    return False


def _required_exam_parts_for_program(program: Program) -> set[str]:
    if not program:
        return set()

    try:
        configs = list(program.exam_parts.filter(is_required=True).order_by('order', 'id'))
    except Exception:
        configs = []
    if configs:
        return {c.part_type for c in configs}

    program_type = getattr(program, 'program_type', None)
    if program_type == Program.ProgramType.TECHNICAL_EXAM:
        return {ProgramExamPartConfig.PartType.PRACTICAL}
    if program_type == Program.ProgramType.SOLAR_POWER_EXAM:
        return {
            ProgramExamPartConfig.PartType.THEORETICAL,
            ProgramExamPartConfig.PartType.PRACTICAL,
        }
    if program_type in {Program.ProgramType.EXAM_ONLY, Program.ProgramType.COURSE_ATTENDANCE}:
        return {ProgramExamPartConfig.PartType.THEORETICAL}
    return set()


def _contractor_can_enter_exam(enrollment: EnrollmentRequest) -> bool:
    if not enrollment.exam_template_id or not enrollment.exam_date:
        return False

    if enrollment.status not in {
        EnrollmentRequest.Status.EXAM_SCHEDULED,
        EnrollmentRequest.Status.EXAM_CONFIRMED,
        EnrollmentRequest.Status.IN_EXAM,
    }:
        return False

    return enrollment.exam_date <= timezone.now()


def _trainer_external_assessment_gate(enrollment: EnrollmentRequest) -> dict:
    required_parts = set(enrollment.required_exam_parts() or [])
    requires_theoretical = ProgramExamPartConfig.PartType.THEORETICAL in required_parts
    requires_external = bool(
        (ProgramExamPartConfig.PartType.PRACTICAL in required_parts)
        or (ProgramExamPartConfig.PartType.PROJECT in required_parts)
    )
    external_only = bool(requires_external and not requires_theoretical)

    is_technical_exam = False
    try:
        is_technical_exam = enrollment.program.program_type == Program.ProgramType.TECHNICAL_EXAM
    except Exception:
        is_technical_exam = False

    can_enter_external_assessment = (
        requires_external
        and enrollment.status in (
            {
                EnrollmentRequest.Status.EXAM_CONFIRMED,
                EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE,
                EnrollmentRequest.Status.FAILED,
            }
            if (is_technical_exam or external_only)
            else {
                EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING,
                EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE,
                EnrollmentRequest.Status.FAILED,
            }
        )
    )
    if requires_theoretical and can_enter_external_assessment and not enrollment.has_passed_theoretical():
        can_enter_external_assessment = False

    locked_by_attempt_limit = _is_locked_after_max_part_attempts(enrollment)
    if locked_by_attempt_limit:
        can_enter_external_assessment = False

    return {
        'required_parts': required_parts,
        'requires_theoretical': requires_theoretical,
        'requires_external': requires_external,
        'external_only': external_only,
        'is_technical_exam': is_technical_exam,
        'is_technical_exam_ui': bool(is_technical_exam or external_only),
        'can_enter_external_assessment': can_enter_external_assessment,
        'locked_by_attempt_limit': locked_by_attempt_limit,
    }


def _options_map_for_questions(question_ids):
    """
    ✅ لا نعتمد على related_name (options) لأننا ما نضمنه.
    نجيب الخيارات مباشرة من ExamOption ونبني map.
    """
    if not question_ids:
        return {}

    order_fields = []
    if _has_field(ExamOption, 'order'):
        order_fields.append('order')
    order_fields.append('id')

    opts = (
        ExamOption.objects
        .filter(question_id__in=question_ids)
        .order_by(*order_fields)
    )
    mp = {}
    for o in opts:
        mp.setdefault(o.question_id, []).append(o)
    return mp


def _has_field(model_cls, field_name: str) -> bool:
    try:
        return any(f.name == field_name for f in model_cls._meta.get_fields())
    except Exception:
        return False


def _scorm_storage() -> FileSystemStorage:
    scorm_dir = os.path.join(settings.MEDIA_ROOT, 'scorm')
    os.makedirs(scorm_dir, exist_ok=True)
    base_url = settings.MEDIA_URL.rstrip('/') + '/scorm/'
    return FileSystemStorage(location=scorm_dir, base_url=base_url)


def _scorm_extracted_storage() -> FileSystemStorage:
    extracted_dir = os.path.join(settings.MEDIA_ROOT, 'scorm_extracted')
    os.makedirs(extracted_dir, exist_ok=True)
    base_url = settings.MEDIA_URL.rstrip('/') + '/scorm_extracted/'
    return FileSystemStorage(location=extracted_dir, base_url=base_url)


def _scorm_metadata_path() -> str:
    return os.path.join(_scorm_storage().location, 'metadata.json')


def _load_scorm_metadata() -> dict:
    path = _scorm_metadata_path()
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except FileNotFoundError:
        return {}
    except Exception:
        return {}


def _save_scorm_metadata(data: dict) -> None:
    path = _scorm_metadata_path()
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        # لا نوقف التطبيق بسبب مشكلة كتابة metadata
        pass


def _safe_extract_zip(zip_abs_path: str, extract_to_dir: str) -> None:
    os.makedirs(extract_to_dir, exist_ok=True)

    with zipfile.ZipFile(zip_abs_path) as zf:
        for member in zf.infolist():
            member_name = member.filename
            if not member_name:
                continue

            # normalize slashes
            member_name = member_name.replace('\\', '/')
            # prevent zip slip
            dest_path = os.path.normpath(os.path.join(extract_to_dir, member_name))
            if not dest_path.startswith(os.path.normpath(extract_to_dir) + os.sep):
                continue

            if member.is_dir() or member_name.endswith('/'):
                os.makedirs(dest_path, exist_ok=True)
                continue

            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            with zf.open(member) as src, open(dest_path, 'wb') as dst:
                dst.write(src.read())


def _find_scorm_launch_relpath(extract_dir: str) -> str | None:
    manifest_path = os.path.join(extract_dir, 'imsmanifest.xml')
    if os.path.exists(manifest_path):
        try:
            tree = ET.parse(manifest_path)
            root = tree.getroot()

            # ابحث عن أي resource فيه href
            for el in root.iter():
                if not str(el.tag).lower().endswith('resource'):
                    continue
                href = el.attrib.get('href')
                if not href:
                    continue
                href = href.replace('\\', '/')
                href = href.lstrip('/')
                candidate = os.path.normpath(os.path.join(extract_dir, href))
                if candidate.startswith(os.path.normpath(extract_dir) + os.sep) and os.path.isfile(candidate):
                    return href
        except Exception:
            pass

    # fallback: index.html at root
    for fallback in ('index.html', 'Index.html', 'INDEX.html'):
        p = os.path.join(extract_dir, fallback)
        if os.path.isfile(p):
            return fallback

    # fallback: first html file
    try:
        for root_dir, _dirs, files in os.walk(extract_dir):
            for fn in sorted(files, key=lambda s: s.lower()):
                if fn.lower().endswith('.html'):
                    abs_path = os.path.join(root_dir, fn)
                    rel = os.path.relpath(abs_path, extract_dir)
                    return rel.replace('\\', '/')
    except Exception:
        return None

    return None


def _ensure_extracted_for_zip(zip_filename: str) -> dict | None:
    """Ensure this ZIP has an extracted folder + launch file; returns metadata entry or None."""
    zip_storage = _scorm_storage()
    if not zip_storage.exists(zip_filename):
        return None

    meta = _load_scorm_metadata()
    entry = meta.get(zip_filename)

    extracted_storage = _scorm_extracted_storage()

    # If we already have an entry, validate it still exists
    if isinstance(entry, dict):
        folder = entry.get('folder')
        launch = entry.get('launch')
        if folder and launch:
            launch_abs = os.path.join(extracted_storage.location, folder, launch)
            if os.path.isfile(launch_abs):
                return entry

    # Create new extraction
    zip_abs = os.path.join(zip_storage.location, zip_filename)
    base = os.path.splitext(os.path.basename(zip_filename))[0]
    folder = f"{base}_{uuid.uuid4().hex[:8]}"
    extract_dir = os.path.join(extracted_storage.location, folder)

    try:
        _safe_extract_zip(zip_abs, extract_dir)
        launch_rel = _find_scorm_launch_relpath(extract_dir)
    except Exception:
        launch_rel = None

    if not launch_rel:
        # لا نكتب entry ناقص
        return None

    meta[zip_filename] = {
        'folder': folder,
        'launch': launch_rel,
        'extracted_at': timezone.now().isoformat(),
    }
    _save_scorm_metadata(meta)
    return meta[zip_filename]


def _get_scorm_entry_for_folder(folder: str) -> tuple[str, dict] | None:
    """Return (zip_filename, entry) if folder matches a known extracted folder."""
    if not folder or '/' in folder or '\\' in folder:
        return None
    meta = _load_scorm_metadata()
    for zip_name, entry in meta.items():
        if isinstance(entry, dict) and entry.get('folder') == folder:
            return zip_name, entry
    return None


@xframe_options_sameorigin
def scorm_player_file_view(request, folder: str, filepath: str):
    """Serve extracted SCORM files with SAMEORIGIN framing enabled."""
    folder = (folder or '').strip()
    filepath = (filepath or '').replace('\\', '/').lstrip('/')

    # ensure folder is known
    found = _get_scorm_entry_for_folder(folder)
    if not found:
        raise Http404('SCORM file not found')

    extracted_storage = _scorm_extracted_storage()
    base_dir = os.path.join(extracted_storage.location, folder)
    base_dir_norm = os.path.normpath(base_dir)

    abs_path = os.path.normpath(os.path.join(base_dir_norm, filepath))
    if not abs_path.startswith(base_dir_norm + os.sep):
        raise Http404('SCORM file not found')

    if not os.path.isfile(abs_path):
        raise Http404('SCORM file not found')

    content_type, _enc = mimetypes.guess_type(abs_path)
    resp = FileResponse(open(abs_path, 'rb'), content_type=content_type or 'application/octet-stream')
    # Ensure browser renders HTML/JS assets in the iframe
    resp['Content-Disposition'] = 'inline'
    return resp


def scorm_player_file_redirect_view(request, folder: str, filepath: str):
    """Redirect legacy trailing-slash player URLs to the canonical no-slash URL."""
    try:
        target = reverse('scorm_player_file', kwargs={'folder': folder, 'filepath': filepath})
    except Exception:
        raise Http404('SCORM file not found')
    return HttpResponsePermanentRedirect(target)


def _list_scorm_packages(include_download_url: bool = True):
    storage = _scorm_storage()
    scorm_dir = storage.location
    extracted_storage = _scorm_extracted_storage()
    meta = _load_scorm_metadata()

    packages = []
    try:
        names = os.listdir(scorm_dir)
    except FileNotFoundError:
        names = []

    for name in sorted(names, key=lambda s: s.lower()):
        if not name.lower().endswith('.zip'):
            continue

        file_path = os.path.join(scorm_dir, name)
        if not os.path.isfile(file_path):
            continue

        try:
            stat = os.stat(file_path)
            size_kb = int(stat.st_size / 1024)
            modified_at = datetime.fromtimestamp(stat.st_mtime)
        except OSError:
            size_kb = None
            modified_at = None

        packages.append({
            'name': name,
            'url': storage.url(name) if include_download_url else None,
            'size_kb': size_kb,
            'modified_at': modified_at,
            'launch_url': None,
        })

        # attach launch url if extracted
        entry = meta.get(name)
        if isinstance(entry, dict) and entry.get('folder') and entry.get('launch'):
            folder = str(entry['folder']).strip()
            launch = str(entry['launch']).replace('\\', '/').lstrip('/')
            launch_abs = os.path.join(extracted_storage.location, folder, launch)
            if os.path.isfile(launch_abs):
                launch_url = extracted_storage.base_url.rstrip('/') + '/' + urllib.parse.quote(folder) + '/' + urllib.parse.quote(launch)
                packages[-1]['launch_url'] = launch_url

    return packages


def _get_scorm_package_or_404(filename: str, include_download_url: bool = True):
    if not filename or '/' in filename or '\\' in filename:
        raise Http404('SCORM package not found')
    if not filename.lower().endswith('.zip'):
        raise Http404('SCORM package not found')

    storage = _scorm_storage()
    if not storage.exists(filename):
        raise Http404('SCORM package not found')

    file_path = os.path.join(storage.location, filename)
    try:
        stat = os.stat(file_path)
        size_kb = int(stat.st_size / 1024)
        modified_at = datetime.fromtimestamp(stat.st_mtime)
    except OSError:
        size_kb = None
        modified_at = None

    extracted_storage = _scorm_extracted_storage()
    entry = _ensure_extracted_for_zip(filename)
    launch_url = None
    if isinstance(entry, dict) and entry.get('folder') and entry.get('launch'):
        folder = str(entry['folder']).strip()
        launch = str(entry['launch']).replace('\\', '/').lstrip('/')
        launch_abs = os.path.join(extracted_storage.location, folder, launch)
        if os.path.isfile(launch_abs):
            launch_url = extracted_storage.base_url.rstrip('/') + '/' + urllib.parse.quote(folder) + '/' + urllib.parse.quote(launch)

    return {
        'name': filename,
        'url': storage.url(filename) if include_download_url else None,
        'size_kb': size_kb,
        'modified_at': modified_at,
        'launch_url': launch_url,
    }


@login_required
def scorm_zip_download_view(request, filename: str):
    """Serve SCORM ZIP downloads only to staff/admins.

    Contractors must not be able to download ZIP packages.
    """
    user = request.user
    if not (getattr(user, 'is_staff', False) or getattr(user, 'is_superuser', False)):
        return HttpResponseForbidden('غير مصرح لك')

    if not filename or '/' in filename or '\\' in filename or not filename.lower().endswith('.zip'):
        raise Http404('SCORM package not found')

    storage = _scorm_storage()
    if not storage.exists(filename):
        raise Http404('SCORM package not found')

    file_path = os.path.join(storage.location, filename)
    resp = FileResponse(open(file_path, 'rb'), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def trainer_scorm_upload_view(request):
    user = request.user
    if not (_is_trainer(user) or _is_super_admin(user)):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    if request.method == 'POST':
        # Trainers can view the list of uploaded SCORM packages, but must not upload.
        if not _is_super_admin(user):
            messages.error(request, 'غير مصرح لك برفع ملفات SCORM')
            return redirect('trainer_scorm_upload')
        return _handle_scorm_upload_post(request, 'trainer_scorm_upload')

    return render(
        request,
        'accounts-templates/trainer-scorm-courses.html',
        {
            'packages': _list_scorm_packages(include_download_url=False),
        }
    )


@login_required
def super_admin_scorm_upload_view(request):
    if not _super_admin_guard(request):
        return redirect('home')

    if request.method == 'POST':
        return _handle_scorm_upload_post(request, 'super_admin_scorm_upload')

    return render(
        request,
        'accounts-templates/superadmin-scorm-upload.html',
        {
            'packages': _list_scorm_packages(),
        }
    )


@login_required
def contractor_scorm_courses_view(request):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/contractor-scorm-courses.html',
        {
            'packages': _list_scorm_packages(include_download_url=False),
        }
    )


@login_required
def contractor_scorm_course_view(request, filename):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    selected = _get_scorm_package_or_404(filename, include_download_url=False)

    # Prefer serving launch via our player route (avoids X-Frame-Options DENY)
    meta = _load_scorm_metadata().get(filename)
    if isinstance(meta, dict) and meta.get('folder') and meta.get('launch'):
        folder = str(meta['folder']).strip()
        launch = str(meta['launch']).replace('\\', '/').lstrip('/')
        try:
            selected['player_url'] = reverse('scorm_player_file', kwargs={'folder': folder, 'filepath': launch})
        except Exception:
            selected['player_url'] = None
    else:
        selected['player_url'] = None

    return render(
        request,
        'accounts-templates/contractor-scorm-course-view.html',
        {
            'packages': _list_scorm_packages(include_download_url=False),
            'selected': selected,
        }
    )


@login_required
@require_POST
@transaction.atomic
def contractor_scorm_check_complete_action(request, filename: str):
    """Checks SCORM completion status (client-reported) and issues a PDF certificate."""
    user = request.user
    if not _is_contractor(user):
        return JsonResponse({'ok': False, 'message': 'غير مصرح'}, status=403)

    selected = _get_scorm_package_or_404(filename)
    course_name = str(selected.get('name') or filename)
    if course_name.lower().endswith('.zip'):
        course_name = course_name[:-4]

    # Client-reported SCORM status values
    lesson_status = (request.POST.get('lesson_status') or '').strip().lower()
    completion_status = (request.POST.get('completion_status') or '').strip().lower()
    success_status = (request.POST.get('success_status') or '').strip().lower()

    completed_values = {'completed', 'passed'}
    completed = False
    if lesson_status in completed_values:
        completed = True
    if completion_status == 'completed':
        completed = True
    if success_status == 'passed':
        completed = True

    if not completed:
        return JsonResponse({
            'ok': False,
            'completed': False,
            'message': 'لم يتم تسجيل إكمال الدورة بعد. أكمل الدورة ثم أعد المحاولة.',
        }, status=200)

    cert, _created = ScormCertificate.objects.get_or_create(
        owner=user,
        scorm_filename=filename,
        defaults={
            'course_name': course_name,
            'verification_code': uuid.uuid4().hex[:12].upper(),
        }
    )

    # keep name fresh
    if not cert.course_name:
        cert.course_name = course_name
        cert.save(update_fields=['course_name'])

    certificate_url = cert.pdf_file.url if cert.pdf_file else None
    certificate_error = None

    if not certificate_url:
        try:
            pdf_bytes = _generate_certificate_pdf_bytes(
                owner_username=_contractor_english_name(user),
                program_name=_english_text_only(course_name, fallback='SCORM Course'),
                verification_code=cert.verification_code,
                issued_at=getattr(cert, 'issued_at', None) or timezone.now(),
                certificate_kind='CERTIFICATE',
            )

            safe_code = (cert.verification_code or uuid.uuid4().hex[:12].upper())
            filename_out = f"scorm_certificate_{user.id}_{safe_code}.pdf"
            cert.pdf_file.save(filename_out, ContentFile(pdf_bytes), save=False)
            cert.save()
            certificate_url = cert.pdf_file.url if cert.pdf_file else None
        except Exception:
            certificate_error = 'تم تسجيل الإكمال، لكن تعذر إنشاء ملف PDF الآن. حاول لاحقًا.'

    return JsonResponse({
        'ok': True,
        'completed': True,
        'certificate_url': certificate_url,
        'certificate_error': certificate_error,
    })


# =========================
# Home
# =========================

def home_view(request):
    if request.user.is_authenticated:
        return redirect(_primary_dashboard_route(request.user))

    return render(request, 'home.html')


# =========================
# Authentication
# =========================

def register_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    default_role = 'business_owner'
    type_hint = (request.GET.get('type') or '').strip().lower()
    if type_hint in {'company', 'business', 'owner'}:
        default_role = 'business_owner'

    if request.method == 'POST':
        post_data = request.POST.copy()
        # دعم صفحة تسجيل المقاول الحالية: لو ما أرسلنا role، اعتبره مقاول
        post_data['role'] = 'business_owner'
        if not post_data.get('phone_number'):
            post_data['phone_number'] = '0500000000'
        if not post_data.get('id_number'):
            post_data['id_number'] = '1000000000'
        if not post_data.get('region'):
            post_data['region'] = 'Central region'
        if not post_data.get('sec_business_line'):
            post_data['sec_business_line'] = 'Facilities Sector'

        form = RegisterForm(post_data)
        form.fields['role'].choices = [('business_owner', 'Business Owner')]
        if form.is_valid():
            username = form.cleaned_data['username']
            email = (form.cleaned_data.get('email') or '').strip()
            full_name_en = (form.cleaned_data.get('full_name_en') or '').strip()
            full_name_ar = (form.cleaned_data.get('full_name_ar') or '').strip()
            password = form.cleaned_data['password']
            role = form.cleaned_data['role']

            if User.objects.filter(username=username).exists():
                messages.error(request, 'اسم المستخدم مستخدم مسبقًا')
                return redirect('register')

            user = User.objects.create_user(username=username, password=password)
            update_fields = []
            if email:
                user.email = email
                update_fields.append('email')
            if full_name_en:
                user.first_name = full_name_en
                update_fields.append('first_name')
            if full_name_ar:
                user.last_name = full_name_ar
                update_fields.append('last_name')
            if update_fields:
                user.save(update_fields=update_fields)

            TrainerProfile.objects.filter(user=user).delete()
            ContractorProfile.objects.filter(user=user).delete()
            BusinessTenant.objects.get_or_create(
                owner=user,
                defaults={
                    'name': (form.cleaned_data.get('company_name') or '').strip() or user.username,
                    'industry': 'Food & Beverage',
                },
            )

            login(request, user)
            messages.success(request, 'تم إنشاء الحساب بنجاح 🎉')
            # حسب طلبك: تحويل مباشر للوحة تحكم المقاول
            return redirect('business_owner_dashboard')
        else:
            messages.error(request, 'تحقق من البيانات المدخلة')
    else:
        form = RegisterForm(initial={'role': default_role})
        form.fields['role'].choices = [('business_owner', 'Business Owner')]

    return render(
        request,
        'accounts-templates/register.html',
        {
            'form': form,
            'default_role': default_role,
        }
    )


def login_view(request):
    if request.user.is_authenticated:
        return redirect(_primary_dashboard_route(request.user))

    login_type = (request.GET.get('type') or '').strip().lower()
    if login_type not in {'individual', 'company'}:
        login_type = 'individual'

    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            if next_url and url_has_allowed_host_and_scheme(
                url=next_url,
                allowed_hosts={request.get_host()},
                require_https=request.is_secure(),
            ):
                return redirect(next_url)

            return redirect(_primary_dashboard_route(user))

        messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')

    return render(
        request,
        'accounts-templates/login.html',
        {
            'login_type': login_type,
            'next': next_url,
        },
    )


def logout_view(request):
    logout(request)
    return redirect('home')


def _business_owner_dashboard_context(request):
    business = _get_owned_business(request.user)
    employees = (
        EmployeeProfile.objects
        .filter(business=business)
        .select_related('user', 'job_title')
        .order_by('user__username')
    )
    courses = business.courses.order_by('title', 'id')
    checklists = business.sop_checklists.prefetch_related('items').order_by('title', 'id')
    job_titles = business.job_titles.order_by('name', 'id')
    course_rules = (
        CourseAssignmentRule.objects
        .filter(business=business)
        .select_related('job_title', 'course')
        .order_by('job_title__name', 'course__title', 'id')
    )
    checklist_rules = (
        SOPChecklistAssignmentRule.objects
        .filter(business=business)
        .select_related('job_title', 'checklist')
        .order_by('job_title__name', 'checklist__title', 'id')
    )
    course_assignment_counts = {
        row['course_id']: row['total']
        for row in (
            CourseAssignment.objects
            .filter(business=business)
            .values('course_id')
            .annotate(total=Count('id'))
        )
    }
    checklist_completion_counts = {
        row['checklist_id']: row['total']
        for row in (
            SOPChecklistCompletion.objects
            .filter(business=business, completed_for=timezone.localdate())
            .values('checklist_id')
            .annotate(total=Count('id'))
        )
    }
    for course in courses:
        course.assignment_total = course_assignment_counts.get(course.id, 0)
    for checklist in checklists:
        checklist.completion_total_today = checklist_completion_counts.get(checklist.id, 0)

    return {
        'business': business,
        'employees': employees,
        'courses': courses,
        'checklists': checklists,
        'job_titles': job_titles,
        'course_rules': course_rules,
        'checklist_rules': checklist_rules,
        'employee_form': BusinessEmployeeCreateForm(business=business),
        'job_title_form': JobTitleForm(),
        'course_form': CourseForm(),
        'course_rule_form': CourseAssignmentRuleForm(business=business),
        'checklist_form': SOPChecklistForm(),
        'checklist_rule_form': SOPChecklistAssignmentRuleForm(business=business),
    }


@login_required
def business_owner_dashboard_view(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/business-owner-dashboard.html',
        _business_owner_dashboard_context(request),
    )


@login_required
def business_owner_employees_view(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/business-owner-employees.html',
        _business_owner_dashboard_context(request),
    )


@login_required
def business_owner_courses_view(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/business-owner-courses.html',
        _business_owner_dashboard_context(request),
    )


@login_required
def business_owner_checklists_view(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/business-owner-checklists.html',
        _business_owner_dashboard_context(request),
    )


@login_required
@require_POST
def business_owner_job_title_create_action(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    business = _get_owned_business(request.user)
    form = JobTitleForm(request.POST)
    if form.is_valid():
        job_title = form.save(commit=False)
        job_title.business = business
        try:
            job_title.save()
            messages.success(request, 'تمت إضافة المسمى الوظيفي')
        except IntegrityError:
            messages.error(request, 'هذا المسمى الوظيفي موجود بالفعل')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('business_owner_employees')


@login_required
@require_POST
@transaction.atomic
def business_owner_employee_create_action(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    business = _get_owned_business(request.user)
    form = BusinessEmployeeCreateForm(request.POST, business=business)
    if not form.is_valid():
        messages.error(request, form.errors.as_text())
        return redirect('business_owner_employees')

    username = form.cleaned_data['username']
    if User.objects.filter(username=username).exists():
        messages.error(request, 'اسم المستخدم مستخدم مسبقاً')
        return redirect('business_owner_employees')

    full_name = form.cleaned_data['full_name']
    first_name, _, last_name = full_name.partition(' ')
    user = User.objects.create_user(
        username=username,
        password=form.cleaned_data['password'],
        email=(form.cleaned_data.get('email') or '').strip(),
        first_name=first_name.strip(),
        last_name=last_name.strip(),
    )
    employee_profile = EmployeeProfile.objects.create(
        user=user,
        business=business,
        job_title=form.cleaned_data.get('job_title'),
        created_by=request.user,
    )
    _provision_course_assignments_for_employee(employee_profile, assigned_by=request.user)
    messages.success(request, 'تم إنشاء حساب الموظف')
    return redirect('business_owner_employees')


@login_required
@require_POST
def business_owner_course_create_action(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    business = _get_owned_business(request.user)
    form = CourseForm(request.POST)
    if form.is_valid():
        course = form.save(commit=False)
        course.business = business
        course.created_by = request.user
        course.save()
        messages.success(request, 'تم إنشاء الدورة')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('business_owner_courses')


@login_required
@require_POST
def business_owner_course_assignment_rule_create_action(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    business = _get_owned_business(request.user)
    form = CourseAssignmentRuleForm(request.POST, business=business)
    if form.is_valid():
        rule = form.save(commit=False)
        rule.business = business
        rule.assigned_by = request.user
        try:
            rule.save()
            _ensure_course_assignments_for_rule(rule)
            messages.success(request, 'تم حفظ قاعدة إسناد الدورة')
        except IntegrityError:
            messages.error(request, 'هذه الدورة مسندة بالفعل لهذا المسمى الوظيفي')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('business_owner_courses')


@login_required
@require_POST
def business_owner_checklist_create_action(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    business = _get_owned_business(request.user)
    form = SOPChecklistForm(request.POST)
    if form.is_valid():
        items = form.cleaned_data['item_lines']
        checklist = form.save(commit=False)
        checklist.business = business
        checklist.created_by = request.user
        checklist.save()
        for index, item_title in enumerate(items, start=1):
            SOPChecklistItem.objects.create(
                checklist=checklist,
                title=item_title,
                order=index,
            )
        messages.success(request, 'تم إنشاء قائمة تشغيلية')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('business_owner_checklists')


@login_required
@require_POST
def business_owner_checklist_assignment_rule_create_action(request):
    if not _business_owner_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    business = _get_owned_business(request.user)
    form = SOPChecklistAssignmentRuleForm(request.POST, business=business)
    if form.is_valid():
        rule = form.save(commit=False)
        rule.business = business
        rule.assigned_by = request.user
        try:
            rule.save()
            messages.success(request, 'تم حفظ قاعدة إسناد قائمة SOP')
        except IntegrityError:
            messages.error(request, 'هذه القائمة مسندة بالفعل لهذا المسمى الوظيفي')
    else:
        messages.error(request, form.errors.as_text())
    return redirect('business_owner_checklists')


@login_required
def employee_dashboard_view(request):
    if not _employee_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/employee-dashboard.html',
        _employee_dashboard_context(request),
    )


def _employee_dashboard_context(request):
    employee_profile = _get_employee_profile(request.user)
    business = employee_profile.business
    today = timezone.localdate()
    # Backfill any missing course assignments so legacy or manually edited employee records stay in sync.
    _provision_course_assignments_for_employee(employee_profile)
    course_assignments = (
        CourseAssignment.objects
        .filter(employee=request.user, business=business)
        .select_related('course')
        .order_by('status', 'course__title', 'id')
    )
    completed_course_count = sum(
        1 for assignment in course_assignments
        if assignment.status == CourseAssignment.Status.COMPLETED
    )
    assigned_checklists = list(_assigned_checklists_queryset(employee_profile))
    today_completions = {
        completion.checklist_id: completion
        for completion in SOPChecklistCompletion.objects.filter(
            business=business,
            employee=request.user,
            completed_for=today,
        ).select_related('checklist')
    }
    recent_checklist_completions = (
        SOPChecklistCompletion.objects
        .filter(business=business, employee=request.user)
        .select_related('checklist')
        .order_by('-completed_for', '-completed_at')[:10]
    )

    return {
        'employee_profile': employee_profile,
        'business': business,
        'course_assignments': course_assignments,
        'completed_course_count': completed_course_count,
        'assigned_checklists': assigned_checklists,
        'today_completions': today_completions,
        'recent_checklist_completions': recent_checklist_completions,
        'today': today,
    }


@login_required
def employee_courses_view(request):
    if not _employee_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/employee-courses.html',
        _employee_dashboard_context(request),
    )


@login_required
def employee_checklists_view(request):
    if not _employee_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    return render(
        request,
        'accounts-templates/employee-checklists.html',
        _employee_dashboard_context(request),
    )


@login_required
@require_POST
def employee_course_complete_action(request, assignment_id: int):
    if not _employee_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    employee_profile = _get_employee_profile(request.user)
    assignment = get_object_or_404(
        CourseAssignment.objects.select_related('course'),
        id=assignment_id,
        employee=request.user,
        business=employee_profile.business,
    )
    assignment.status = CourseAssignment.Status.COMPLETED
    assignment.completed_at = timezone.now()
    assignment.save(update_fields=['status', 'completed_at'])
    messages.success(request, f'تم إكمال الدورة: {assignment.course.title}')
    return redirect('employee_courses')


@login_required
@require_POST
@transaction.atomic
def employee_checklist_complete_action(request, checklist_id: int):
    if not _employee_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    employee_profile = _get_employee_profile(request.user)
    checklist = get_object_or_404(
        _assigned_checklists_queryset(employee_profile),
        id=checklist_id,
    )
    items = list(checklist.items.all())
    selected_item_ids = {int(value) for value in request.POST.getlist('item_ids') if str(value).isdigit()}
    expected_item_ids = {item.id for item in items}
    if expected_item_ids and selected_item_ids != expected_item_ids:
        messages.error(request, 'يجب تحديد جميع عناصر قائمة SOP قبل الإكمال')
        return redirect('employee_checklists')

    completion, _created = SOPChecklistCompletion.objects.get_or_create(
        business=employee_profile.business,
        checklist=checklist,
        employee=request.user,
        completed_for=timezone.localdate(),
        defaults={'notes': (request.POST.get('notes') or '').strip()},
    )
    if not _created:
        completion.notes = (request.POST.get('notes') or '').strip()
        completion.save(update_fields=['notes'])

    for item in items:
        SOPChecklistItemCompletion.objects.update_or_create(
            completion=completion,
            item=item,
            defaults={'is_checked': True},
        )

    messages.success(request, f'تم إكمال قائمة SOP: {checklist.title}')
    return redirect('employee_checklists')


# =========================
# Dashboards
# =========================

@login_required
def super_admin_dashboard_view(request):
    if not _super_admin_guard(request):
        return redirect('home')

    status_rows = (
        EnrollmentRequest.objects
        .values('status')
        .annotate(c=Count('id'))
    )
    status_counts = {row['status']: row['c'] for row in status_rows}

    ctx = {
        'contractors_count': ContractorProfile.objects.count(),
        'trainers_count': TrainerProfile.objects.count(),
        'programs_count': Program.objects.count(),
        'active_programs_count': Program.objects.filter(is_active=True).count(),
        'requests_total': EnrollmentRequest.objects.count(),
        'requests_new': status_counts.get(EnrollmentRequest.Status.NEW_REQUEST, 0),
        'requests_invoice': (
            status_counts.get(EnrollmentRequest.Status.INVOICE_ISSUED, 0)
            + status_counts.get(EnrollmentRequest.Status.PAYMENT_VERIFICATION, 0)
        ),
        'requests_waiting_schedule': status_counts.get(EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING, 0),
        'requests_in_exam': status_counts.get(EnrollmentRequest.Status.IN_EXAM, 0),
        'requests_certified': status_counts.get(EnrollmentRequest.Status.CERTIFIED, 0),
        'certificates_total': Certificate.objects.count(),
        'scorm_certificates_total': ScormCertificate.objects.count(),
    }

    return render(request, 'accounts-templates/superadmin-dashboard.html', ctx)


# =========================
# Super Admin – Programs
# =========================


@login_required
def super_admin_programs_list_view(request):
    if not _super_admin_guard(request):
        return redirect('home')

    q = (request.GET.get('q') or '').strip()
    active = (request.GET.get('active') or '').strip().lower()
    program_type = (request.GET.get('program_type') or '').strip()

    all_programs = Program.objects.all()
    choice_map = dict(Program.ProgramType.choices)
    existing_types = set(
        all_programs.values_list('program_type', flat=True).distinct()
    )
    program_type_choices = [
        (code, label)
        for code, label in Program.ProgramType.choices
        if code in existing_types
    ]
    if program_type and program_type not in {code for code, _label in program_type_choices}:
        program_type_choices.append((program_type, choice_map.get(program_type, program_type)))

    programs = all_programs.order_by('-id')
    if q:
        programs = programs.filter(Q(title__icontains=q) | Q(description__icontains=q))
    if program_type:
        programs = programs.filter(program_type=program_type)
    if active in {'1', 'true', 'yes', 'on'}:
        programs = programs.filter(is_active=True)
    elif active in {'0', 'false', 'no', 'off'}:
        programs = programs.filter(is_active=False)

    return render(
        request,
        'accounts-templates/superadmin-programs.html',
        {
            'programs': programs,
            'q': q,
            'active': active,
            'program_type': program_type,
            'program_type_choices': program_type_choices,
        }
    )


@login_required
def super_admin_program_create_view(request):
    if not _super_admin_guard(request):
        return redirect('home')

    if request.method == 'POST':
        form = ProgramForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم إنشاء البرنامج بنجاح')
            return redirect('super_admin_programs')
        messages.error(request, 'تحقق من البيانات المدخلة')
    else:
        form = ProgramForm()

    return render(
        request,
        'accounts-templates/superadmin-program-form.html',
        {
            'form': form,
            'mode': 'create',
            'subcategory_map_json': form.subcategory_map_json,
            'tertiary_map_json': form.tertiary_map_json,
        }
    )


@login_required
def super_admin_program_edit_view(request, program_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    program = get_object_or_404(Program, id=program_id)

    if request.method == 'POST':
        form = ProgramForm(request.POST, instance=program)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث البرنامج بنجاح')
            return redirect('super_admin_programs')
        messages.error(request, 'تحقق من البيانات المدخلة')
    else:
        form = ProgramForm(instance=program)

    return render(
        request,
        'accounts-templates/superadmin-program-form.html',
        {
            'form': form,
            'program': program,
            'mode': 'edit',
            'subcategory_map_json': form.subcategory_map_json,
            'tertiary_map_json': form.tertiary_map_json,
        }
    )


@login_required
@require_POST
def super_admin_program_delete_action(request, program_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    program = get_object_or_404(Program, id=program_id)
    linked_requests_count = EnrollmentRequest.objects.filter(program=program).count()
    if linked_requests_count > 0:
        messages.error(
            request,
            'لا يمكن حذف البرنامج لوجود طلبات تسجيل مرتبطة به.'
        )
        return redirect('super_admin_program_edit', program_id=program.id)

    program_title = program.title
    program.delete()
    messages.success(request, f'تم حذف البرنامج "{program_title}" بنجاح')
    return redirect('super_admin_programs')


# =========================
# Super Admin – Program Grading Config
# =========================


@login_required
@require_http_methods(['GET', 'POST'])
def super_admin_program_grading_config_view(request, program_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    program = get_object_or_404(Program, id=program_id)

    # Load existing config rows (at most 1 per part type)
    existing = {c.part_type: c for c in program.exam_parts.all()}

    from .forms import ProgramExamPartsConfigForm
    from training.models import ProgramExamPartConfig

    if request.method == 'POST':
        form = ProgramExamPartsConfigForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            desired = []
            order = 1

            def _push(part_type: str, passing_key: str):
                nonlocal order
                passing = int(cd.get(passing_key))
                desired.append((part_type, passing, order))
                order += 1

            if cd.get('include_theoretical'):
                _push(ProgramExamPartConfig.PartType.THEORETICAL, 'theoretical_passing')
            if cd.get('include_practical'):
                _push(ProgramExamPartConfig.PartType.PRACTICAL, 'practical_passing')
            if cd.get('include_project'):
                _push(ProgramExamPartConfig.PartType.PROJECT, 'project_passing')

            desired_types = {t for t, _p, _o in desired}

            # Delete removed parts
            ProgramExamPartConfig.objects.filter(program=program).exclude(part_type__in=desired_types).delete()

            # Upsert desired parts
            for part_type, passing, order in desired:
                obj = existing.get(part_type)
                if obj is None:
                    ProgramExamPartConfig.objects.create(
                        program=program,
                        part_type=part_type,
                        passing_grade_percent=passing,
                        order=order,
                        is_required=True,
                    )
                else:
                    updates = {}
                    if int(getattr(obj, 'passing_grade_percent', 0) or 0) != passing:
                        updates['passing_grade_percent'] = passing
                    if int(getattr(obj, 'order', 0) or 0) != order:
                        updates['order'] = order
                    if not getattr(obj, 'is_required', True):
                        updates['is_required'] = True
                    if updates:
                        for k, v in updates.items():
                            setattr(obj, k, v)
                        obj.save(update_fields=list(updates.keys()))

            messages.success(request, 'تم حفظ إعدادات أجزاء الاختبار')
            return redirect('super_admin_program_edit', program_id=program.id)

        messages.error(request, 'تحقق من البيانات المدخلة')
    else:
        initial = {
            'include_theoretical': ProgramExamPartConfig.PartType.THEORETICAL in existing,
            'theoretical_passing': getattr(existing.get(ProgramExamPartConfig.PartType.THEORETICAL), 'passing_grade_percent', None),
            'include_practical': ProgramExamPartConfig.PartType.PRACTICAL in existing,
            'practical_passing': getattr(existing.get(ProgramExamPartConfig.PartType.PRACTICAL), 'passing_grade_percent', None),
            'include_project': ProgramExamPartConfig.PartType.PROJECT in existing,
            'project_passing': getattr(existing.get(ProgramExamPartConfig.PartType.PROJECT), 'passing_grade_percent', None),
        }
        form = ProgramExamPartsConfigForm(initial=initial)

    return render(
        request,
        'accounts-templates/superadmin-program-grading.html',
        {
            'program': program,
            'form': form,
        }
    )


# =========================
# Super Admin – Users & Roles
# =========================


@login_required
def super_admin_users_list_view(request):
    if not _super_admin_guard(request):
        return redirect('home')

    q = (request.GET.get('q') or '').strip()
    role = (request.GET.get('role') or '').strip().lower()

    users = User.objects.all().order_by('-id')
    if q:
        users = users.filter(Q(username__icontains=q) | Q(email__icontains=q))

    # Enforce invariant: if user is a trainer, they must not keep contractor profile.
    trainer_user_ids_all = list(TrainerProfile.objects.values_list('user_id', flat=True))
    if trainer_user_ids_all:
        ContractorProfile.objects.filter(user_id__in=trainer_user_ids_all).delete()

    if role == 'trainer':
        users = users.filter(id__in=TrainerProfile.objects.values('user_id'))
    elif role == 'contractor':
        # Contractors are users that have contractor profile and are not trainers.
        users = users.filter(id__in=ContractorProfile.objects.filter(is_training_coordinator=False).values('user_id'))
    elif role == 'coordinator':
        users = users.filter(id__in=ContractorProfile.objects.filter(is_training_coordinator=True).values('user_id'))
    elif role == 'super':
        users = users.filter(is_superuser=True)

    trainer_user_ids = set(TrainerProfile.objects.values_list('user_id', flat=True))
    contractor_user_ids = set(ContractorProfile.objects.values_list('user_id', flat=True))
    coordinator_user_ids = set(
        ContractorProfile.objects.filter(is_training_coordinator=True).values_list('user_id', flat=True)
    )

    return render(
        request,
        'accounts-templates/superadmin-users.html',
        {
            'users': users,
            'q': q,
            'role': role,
            'trainer_user_ids': trainer_user_ids,
            'contractor_user_ids': contractor_user_ids,
            'coordinator_user_ids': coordinator_user_ids,
        }
    )


@login_required
@require_POST
@transaction.atomic
def super_admin_set_user_role_action(request, user_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    target = get_object_or_404(User, id=user_id)
    selected_role = (request.POST.get('role') or '').strip().lower()
    allowed_roles = {'super', 'trainer', 'coordinator', 'contractor'}

    if selected_role not in allowed_roles:
        messages.error(request, 'الدور المحدد غير صالح')
        return redirect('super_admin_users')

    # Safety: prevent current super admin from demoting self accidentally.
    if target.id == request.user.id and selected_role != 'super':
        messages.error(request, 'لا يمكنك إزالة صلاحية Super Admin من حسابك الحالي')
        return redirect('super_admin_users')

    if selected_role == 'super':
        TrainerProfile.objects.filter(user=target).delete()
        ContractorProfile.objects.filter(user=target).delete()
        if not target.is_superuser or not target.is_staff:
            target.is_superuser = True
            target.is_staff = True
            target.save(update_fields=['is_superuser', 'is_staff'])
        messages.success(request, 'تم حفظ الدور: Super Admin')
        return redirect('super_admin_users')

    if target.is_superuser or target.is_staff:
        target.is_superuser = False
        target.is_staff = False
        target.save(update_fields=['is_superuser', 'is_staff'])

    if selected_role == 'trainer':
        ContractorProfile.objects.filter(user=target).delete()
        TrainerProfile.objects.get_or_create(
            user=target,
            defaults={'specialization': 'غير محدد'},
        )
        messages.success(request, 'تم حفظ الدور: مدرب')
        return redirect('super_admin_users')

    # Contractor/coordinator roles rely on existing contractor profile fields.
    profile = ContractorProfile.objects.filter(user=target).first()
    if not profile:
        messages.error(request, 'لا يمكن تعيين هذا الدور لأن ملف المقاول غير موجود لهذا المستخدم')
        return redirect('super_admin_users')

    TrainerProfile.objects.filter(user=target).delete()
    profile.is_training_coordinator = (selected_role == 'coordinator')
    profile.save(update_fields=['is_training_coordinator'])

    if selected_role == 'coordinator':
        messages.success(request, 'تم حفظ الدور: منسق تدريب')
    else:
        messages.success(request, 'تم حفظ الدور: مقاول')
    return redirect('super_admin_users')


@login_required
@require_POST
def super_admin_grant_training_coordinator_role_action(request, user_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    target = get_object_or_404(User, id=user_id)
    if getattr(target, 'is_superuser', False):
        messages.error(request, 'لا يمكن تعديل صلاحيات Super Admin')
        return redirect('super_admin_users')

    if TrainerProfile.objects.filter(user=target).exists():
        messages.error(request, 'هذا المستخدم لديه دور مدرب، لا يمكن منحه دور منسق تدريب')
        return redirect('super_admin_users')

    prof = ContractorProfile.objects.filter(user=target).first()
    if not prof:
        messages.error(request, 'يجب أن يكون المستخدم مقاولاً أولاً (ملف مقاول موجود)')
        return redirect('super_admin_users')

    if prof.is_training_coordinator:
        messages.info(request, 'المستخدم لديه دور منسق تدريب بالفعل')
        return redirect('super_admin_users')

    prof.is_training_coordinator = True
    prof.save(update_fields=['is_training_coordinator'])
    messages.success(request, 'تم منح المستخدم دور منسق تدريب')
    return redirect('super_admin_users')


@login_required
@require_POST
def super_admin_remove_training_coordinator_role_action(request, user_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    target = get_object_or_404(User, id=user_id)
    prof = ContractorProfile.objects.filter(user=target).first()
    if not prof:
        messages.info(request, 'لا يوجد ملف مقاول لهذا المستخدم')
        return redirect('super_admin_users')

    if not prof.is_training_coordinator:
        messages.info(request, 'المستخدم ليس منسق تدريب')
        return redirect('super_admin_users')

    prof.is_training_coordinator = False
    prof.save(update_fields=['is_training_coordinator'])
    messages.success(request, 'تم إزالة دور منسق التدريب')
    return redirect('super_admin_users')


@login_required
@require_POST
def super_admin_grant_trainer_role_action(request, user_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    target = get_object_or_404(User, id=user_id)
    if TrainerProfile.objects.filter(user=target).exists():
        messages.info(request, 'المستخدم لديه دور مدرب بالفعل')
        return redirect('super_admin_users')

    # specialization is required in the model; we use a safe default.
    specialization = (request.POST.get('specialization') or '').strip() or 'غير محدد'

    # Ensure single role: switching to trainer removes contractor profile.
    ContractorProfile.objects.filter(user=target).delete()
    TrainerProfile.objects.create(user=target, specialization=specialization)
    messages.success(request, 'تم منح المستخدم دور مدرب بنجاح')
    return redirect('super_admin_users')


@login_required
@require_POST
def super_admin_remove_trainer_role_action(request, user_id: int):
    if not _super_admin_guard(request):
        return redirect('home')

    target = get_object_or_404(User, id=user_id)
    TrainerProfile.objects.filter(user=target).delete()
    messages.success(request, 'تم إزالة دور المدرب')
    return redirect('super_admin_users')


# =========================
# Super Admin – Enrollments
# =========================


@login_required
@require_http_methods(['GET', 'POST'])
def super_admin_enroll_user_view(request):
    if not _super_admin_guard(request):
        return redirect('home')

    selected_user_id = (request.POST.get('user_id') or request.GET.get('user_id') or '').strip()
    selected_program_id = (request.POST.get('program_id') or request.GET.get('program_id') or '').strip()

    selected_user = None
    if selected_user_id.isdigit():
        selected_user = User.objects.filter(id=int(selected_user_id)).first()

    selected_program = None
    if selected_program_id.isdigit():
        selected_program = Program.objects.filter(id=int(selected_program_id)).first()

    contractor_documents = []
    if selected_user is not None:
        contractor_documents = list(
            ContractorDocument.objects.filter(owner=selected_user).order_by('-uploaded_at', '-id')
        )

    if request.method == 'POST':
        if not selected_user or not selected_program:
            messages.error(request, 'اختر المستخدم والبرنامج')
            return redirect('super_admin_enroll_user')

        if not ContractorProfile.objects.filter(user=selected_user).exists():
            messages.error(request, 'هذا المستخدم ليس لديه ملف مقاول، لا يمكن تسجيله كمتدرب في البرامج')
            return redirect('super_admin_enroll_user')

        doc_ids = request.POST.getlist('document_ids')
        docs = ContractorDocument.objects.filter(owner=selected_user, id__in=doc_ids)

        try:
            terminal_statuses = {
                EnrollmentRequest.Status.REJECTED,
                EnrollmentRequest.Status.FAILED,
                EnrollmentRequest.Status.CERTIFIED,
                EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD,
            }
            enrollment = (
                EnrollmentRequest.objects
                .filter(contractor=selected_user, program=selected_program)
                .exclude(status__in=terminal_statuses)
                .order_by('-created_at', '-id')
                .first()
            )
            created = False
            if enrollment is None:
                enrollment = EnrollmentRequest.objects.create(
                    contractor=selected_user,
                    program=selected_program,
                )
                created = True
        except Exception:
            messages.error(request, 'تعذر إنشاء الطلب')
            return redirect('super_admin_enroll_user')

        if created:
            try:
                enrollment.start_workflow()
                enrollment.save(update_fields=['status'])
            except Exception:
                pass

        try:
            if docs.exists():
                enrollment.supporting_documents.set(docs)
                enrollment.snapshot_supporting_documents(docs)
        except Exception:
            pass

        if created:
            messages.success(request, 'تم تسجيل المستخدم في البرنامج بنجاح')
        else:
            messages.info(request, 'يوجد طلب تسجيل سابق لهذا المستخدم في هذا البرنامج (تم تحديث المرفقات إن وُجدت)')

        return redirect('super_admin_enroll_user')

    # Lists for selects
    users_q = User.objects.all().order_by('username')
    programs_q = Program.objects.all().order_by('-is_active', 'title', '-id')

    return render(
        request,
        'accounts-templates/superadmin-enroll.html',
        {
            'users': users_q,
            'programs': programs_q,
            'selected_user': selected_user,
            'selected_program': selected_program,
            'contractor_documents': contractor_documents,
        }
    )

@login_required
def contractor_dashboard_view(request):
    user = request.user
    if _ensure_legacy_business_owner(user):
        return redirect('business_owner_dashboard')
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    contractor_profile = ContractorProfile.objects.select_related('user').get(user=user)

    contractor_documents = (
        ContractorDocument.objects
        .filter(owner=user)
        .order_by('-uploaded_at', '-id')
    )

    latest_request = (
        EnrollmentRequest.objects
        .filter(contractor=user)
        .select_related('program')
        .order_by('-created_at', '-id')
        .first()
    )

    certificates = list(
        Certificate.objects
        .filter(owner=user)
        .select_related('program', 'enrollment')
        .order_by('-issued_at', '-id')
    )

    scorm_certificates = list(
        ScormCertificate.objects
        .filter(owner=user)
        .order_by('-issued_at', '-id')
    )

    # ✅ Best-effort: regenerate missing PDFs if dependencies are available
    # (prevents users who already passed from losing download links)
    try:
        regen_count = 0
        for cert in certificates:
            if regen_count >= 3:
                break
            if getattr(cert, 'pdf_file', None):
                continue
            if getattr(cert, 'pdf_file', None) and getattr(cert.pdf_file, 'name', None):
                continue

            if not getattr(cert, 'verification_code', None):
                cert.verification_code = uuid.uuid4().hex[:12].upper()
                cert.save(update_fields=['verification_code'])

            pdf_bytes = _generate_certificate_pdf_bytes(
                owner_username=_contractor_english_name(getattr(cert, 'owner', user)),
                program_name=_program_english_name(getattr(cert, 'program', None)),
                verification_code=cert.verification_code,
                issued_at=getattr(cert, 'issued_at', None) or timezone.now(),
                certificate_kind=getattr(cert, 'certificate_type', None) or 'CERTIFICATE',
            )

            filename = f"certificate_{getattr(cert, 'enrollment_id', 'x')}_{cert.verification_code}.pdf"
            cert.pdf_file.save(filename, ContentFile(pdf_bytes), save=False)
            cert.save()
            regen_count += 1

        regen_scorm_count = 0
        for sc in scorm_certificates:
            if regen_scorm_count >= 3:
                break
            if getattr(sc, 'pdf_file', None) and getattr(sc.pdf_file, 'name', None):
                continue
            if not getattr(sc, 'verification_code', None):
                sc.verification_code = uuid.uuid4().hex[:12].upper()
                sc.save(update_fields=['verification_code'])

            pdf_bytes = _generate_certificate_pdf_bytes(
                owner_username=_contractor_english_name(getattr(sc, 'owner', user)),
                program_name=_english_text_only(getattr(sc, 'course_name', None) or 'SCORM', fallback='SCORM Course'),
                verification_code=sc.verification_code,
                issued_at=getattr(sc, 'issued_at', None) or timezone.now(),
                certificate_kind='CERTIFICATE',
            )
            filename = f"scorm_certificate_{user.id}_{sc.verification_code}.pdf"
            sc.pdf_file.save(filename, ContentFile(pdf_bytes), save=False)
            sc.save()
            regen_scorm_count += 1
    except Exception:
        # keep dashboard stable even if PDF generation fails
        pass

    return render(
        request,
        'accounts-templates/contractor-dashboard.html',
        {
            'contractor_profile': contractor_profile,
            'contractor_documents': contractor_documents,
            'certificates': certificates,
            'scorm_certificates': scorm_certificates,
            'latest_request': latest_request,
            'latest_request_status_display': latest_request.get_status_display() if latest_request else 'لا توجد طلبات بعد',
            'latest_request_program_name': _safe_program_name(getattr(latest_request, 'program', None)) if latest_request else None,
        }
    )


@login_required
@require_POST
def contractor_document_upload_action(request):
    user = request.user
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    pdf_file = request.FILES.get('pdf_file')
    title = (request.POST.get('title') or '').strip()

    if not pdf_file:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': 'اختر ملف PDF للرفع'}, status=400)
        messages.error(request, 'اختر ملف PDF للرفع')
        return redirect('contractor_dashboard')

    name = (getattr(pdf_file, 'name', '') or '').lower()
    if not name.endswith('.pdf'):
        if is_ajax:
            return JsonResponse({'ok': False, 'error': 'الملف المسموح به فقط: PDF'}, status=400)
        messages.error(request, 'الملف المسموح به فقط: PDF')
        return redirect('contractor_dashboard')

    content_type = (getattr(pdf_file, 'content_type', '') or '').lower()
    if content_type and 'pdf' not in content_type:
        if is_ajax:
            return JsonResponse({'ok': False, 'error': 'الملف المسموح به فقط: PDF'}, status=400)
        messages.error(request, 'الملف المسموح به فقط: PDF')
        return redirect('contractor_dashboard')

    ContractorDocument.objects.create(
        owner=user,
        title=title,
        pdf_file=pdf_file,
    )

    messages.success(request, 'تم رفع الملف بنجاح')
    return redirect('contractor_dashboard')


@login_required
@require_POST
def contractor_document_delete_action(request, doc_id: int):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    doc = get_object_or_404(ContractorDocument, id=doc_id, owner=user)
    linked_enrollments = list(doc.enrollments.all())
    for enrollment in linked_enrollments:
        try:
            enrollment.snapshot_supporting_documents([doc])
        except Exception:
            pass

    try:
        if doc.pdf_file:
            doc.pdf_file.delete(save=False)
    except Exception:
        pass

    doc.delete()
    messages.success(request, 'تم حذف الملف')
    return redirect('contractor_dashboard')


@login_required
def trainer_dashboard_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    # Trainer dashboard page removed; keep route for backward compatibility.
    return redirect('trainer_requests')


# =========================
# Training Coordinator – Dashboard / Contractors
# =========================


@login_required
def training_coordinator_dashboard_view(request):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    user = request.user
    coordinator_profile = ContractorProfile.objects.filter(user=user).select_related('user').first()

    managed_contractors = (
        ContractorProfile.objects
        .filter(registered_by=user)
        .select_related('user')
        .order_by('-id')
    )

    managed_requests = (
        EnrollmentRequest.objects
        .filter(contractor__contractorprofile__registered_by=user)
        .select_related('program', 'contractor', 'trainer', 'exam_template')
        .order_by('-created_at')
    )

    recent_requests = list(managed_requests[:15])

    return render(
        request,
        'accounts-templates/training-coordinator-dashboard.html',
        {
            'coordinator_profile': coordinator_profile,
            'managed_contractors': managed_contractors,
            'managed_contractors_count': managed_contractors.count(),
            'managed_requests_count': managed_requests.count(),
            'recent_requests': recent_requests,
        }
    )


@login_required
def training_coordinator_contractors_list_view(request):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالدخول')
        return redirect('home')

    user = request.user
    contractor_id = (request.GET.get('contractor_id') or '').strip()
    username_query = (request.GET.get('username') or '').strip()
    region_query = (request.GET.get('region') or '').strip()
    managed_contractors = (
        ContractorProfile.objects
        .filter(registered_by=user)
        .select_related('user')
        .prefetch_related('user__contractor_documents')
        .order_by('-id')
    )
    managed_contractors_total_count = managed_contractors.count()
    available_regions = list(
        managed_contractors
        .exclude(region__isnull=True)
        .exclude(region__exact='')
        .values_list('region', flat=True)
        .distinct()
        .order_by('region')
    )

    if contractor_id:
        managed_contractors = managed_contractors.filter(id_number__icontains=contractor_id)
    if username_query:
        managed_contractors = managed_contractors.filter(user__username__icontains=username_query)
    if region_query:
        managed_contractors = managed_contractors.filter(region=region_query)

    available_programs = Program.objects.filter(is_active=True).order_by('title', 'id')
    has_filters = bool(contractor_id or username_query or region_query)

    return render(
        request,
        'accounts-templates/training-coordinator-contractor-list.html',
        {
            'managed_contractors': managed_contractors,
            'managed_contractors_count': managed_contractors.count(),
            'managed_contractors_total_count': managed_contractors_total_count,
            'contractor_id_query': contractor_id,
            'username_query': username_query,
            'region_query': region_query,
            'available_regions': available_regions,
            'has_filters': has_filters,
            'available_programs': available_programs,
        }
    )

@login_required
@require_POST
def training_coordinator_register_existing_contractor_action(request, contractor_user_id: int):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    coordinator = request.user
    contractor_user, _contractor_profile = _get_managed_contractor_or_404(coordinator, contractor_user_id)

    program_id_raw = (request.POST.get('program_id') or '').strip()
    document_id_raw = (request.POST.get('document_id') or '').strip()

    if not program_id_raw:
        messages.error(request, 'اختر برنامج تدريبي')
        return redirect('training_coordinator_contractors_list')

    if not document_id_raw:
        messages.error(request, 'اختر ملف PDF واحد على الأقل')
        return redirect('training_coordinator_contractors_list')

    try:
        program_id = int(program_id_raw)
        document_id = int(document_id_raw)
    except ValueError:
        messages.error(request, 'بيانات غير صحيحة')
        return redirect('training_coordinator_contractors_list')

    program = get_object_or_404(Program, id=program_id, is_active=True)
    selected_document = get_object_or_404(
        ContractorDocument,
        id=document_id,
        owner=contractor_user,
    )

    terminal_statuses = {
        EnrollmentRequest.Status.REJECTED,
        EnrollmentRequest.Status.FAILED,
        EnrollmentRequest.Status.CERTIFIED,
        EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD,
    }
    active_existing = (
        EnrollmentRequest.objects
        .filter(contractor=contractor_user, program=program)
        .exclude(status__in=terminal_statuses)
        .order_by('-created_at', '-id')
        .first()
    )
    if active_existing is not None:
        messages.info(
            request,
            f'للمقاول طلب قائم بالفعل. الحالة الحالية: {active_existing.get_status_display()}'
        )
        return redirect('training_coordinator_contractors_list')

    enrollment = EnrollmentRequest.objects.create(
        contractor=contractor_user,
        program=program,
    )
    enrollment.start_workflow()
    enrollment.save(update_fields=['status'])
    enrollment.supporting_documents.add(selected_document)
    enrollment.snapshot_supporting_documents([selected_document])

    messages.success(request, 'تم بدء إجراءات تسجيل المقاول في البرنامج بنجاح')
    return redirect('training_coordinator_contractors_list')

@login_required
@require_http_methods(['GET', 'POST'])
def training_coordinator_register_contractor_view(request):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    coordinator = request.user

    if request.method == 'POST':
        form = TrainingCoordinatorRegisterContractorForm(request.POST, request.FILES)
        if form.is_valid():
            username = form.cleaned_data['username']
            email = (form.cleaned_data.get('email') or '').strip()
            password = form.cleaned_data['password']
            selected_program = form.cleaned_data.get('program')

            pdf_files = form.cleaned_data.get('pdf_files') or []

            # Validate PDFs early to avoid partial user creation
            for f in pdf_files:
                if not f:
                    continue
                name = (getattr(f, 'name', '') or '').lower()
                if not name.endswith('.pdf'):
                    messages.error(request, 'الملف المسموح به فقط: PDF')
                    return redirect('training_coordinator_register_contractor')
                content_type = (getattr(f, 'content_type', '') or '').lower()
                if content_type and 'pdf' not in content_type:
                    messages.error(request, 'الملف المسموح به فقط: PDF')
                    return redirect('training_coordinator_register_contractor')

            if User.objects.filter(username=username).exists():
                messages.error(request, 'اسم المستخدم مستخدم مسبقًا')
                return redirect('training_coordinator_register_contractor')

            user = User.objects.create_user(username=username, password=password)
            if email:
                user.email = email
                user.save(update_fields=['email'])

            # Ensure user is a contractor (not a trainer)
            TrainerProfile.objects.filter(user=user).delete()

            ContractorProfile.objects.create(
                user=user,
                company_name=form.cleaned_data['company_name'],
                phone_number=form.cleaned_data['phone_number'],
                id_number=form.cleaned_data.get('id_number') or None,
                region=form.cleaned_data.get('region') or None,
                registered_by=coordinator,
            )

            # Save supporting documents (optional)
            created_docs = []
            for f in pdf_files:
                if not f:
                    continue
                title = (getattr(f, 'name', '') or '').split('/')[-1]
                created_docs.append(
                    ContractorDocument.objects.create(
                        owner=user,
                        title=title,
                        pdf_file=f,
                    )
                )

            # Create enrollment request (optional)
            if selected_program is not None:
                try:
                    terminal_statuses = {
                        EnrollmentRequest.Status.REJECTED,
                        EnrollmentRequest.Status.FAILED,
                        EnrollmentRequest.Status.CERTIFIED,
                        EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD,
                    }
                    enrollment = (
                        EnrollmentRequest.objects
                        .filter(contractor=user, program=selected_program)
                        .exclude(status__in=terminal_statuses)
                        .order_by('-created_at', '-id')
                        .first()
                    )
                    if enrollment is None:
                        enrollment = EnrollmentRequest.objects.create(
                            contractor=user,
                            program=selected_program,
                        )
                    if created_docs:
                        enrollment.supporting_documents.add(*created_docs)
                        enrollment.snapshot_supporting_documents(created_docs)
                except Exception:
                    messages.warning(request, 'تم إنشاء حساب المقاول، لكن تعذر إنشاء طلب البرنامج الآن')

            if selected_program is not None:
                messages.success(request, 'تم تسجيل مقاول جديد وإنشاء طلب برنامج بنجاح')
            else:
                messages.success(request, 'تم تسجيل مقاول جديد بنجاح')
            return redirect('training_coordinator_dashboard')

        messages.error(request, 'تحقق من البيانات المدخلة')
    else:
        form = TrainingCoordinatorRegisterContractorForm()

    return render(
        request,
        'accounts-templates/training-coordinator-register-contractor.html',
        {'form': form}
    )


def _get_managed_contractor_or_404(coordinator_user, contractor_user_id: int):
    contractor_user = get_object_or_404(User, id=contractor_user_id)
    profile = get_object_or_404(
        ContractorProfile.objects.select_related('user'),
        user=contractor_user,
        registered_by=coordinator_user,
    )
    return contractor_user, profile


@login_required
def training_coordinator_requests_list_view(request):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    user = request.user
    requests_qs = (
        EnrollmentRequest.objects
        .filter(contractor__contractorprofile__registered_by=user)
        .select_related('program', 'trainer', 'contractor', 'exam_template')
        .order_by('-created_at')
    )

    return render(
        request,
        'accounts-templates/contractor-requests.html',
        {
            'requests': requests_qs,
            'is_coordinator': True,
        }
    )


@login_required
def training_coordinator_contractor_documents_view(request, contractor_user_id: int):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    coordinator = request.user
    contractor_user, contractor_profile = _get_managed_contractor_or_404(coordinator, contractor_user_id)

    docs = ContractorDocument.objects.filter(owner=contractor_user).order_by('-uploaded_at', '-id')

    return render(
        request,
        'accounts-templates/training-coordinator-contractor-documents.html',
        {
            'contractor_user': contractor_user,
            'contractor_profile': contractor_profile,
            'docs': docs,
        }
    )


@login_required
@require_POST
def training_coordinator_contractor_document_upload_action(request, contractor_user_id: int):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    coordinator = request.user
    contractor_user, _contractor_profile = _get_managed_contractor_or_404(coordinator, contractor_user_id)

    pdf_file = request.FILES.get('pdf_file')
    title = (request.POST.get('title') or '').strip()

    if not pdf_file:
        messages.error(request, 'اختر ملف PDF للرفع')
        return redirect('training_coordinator_contractor_documents', contractor_user_id=contractor_user_id)

    name = (getattr(pdf_file, 'name', '') or '').lower()
    if not name.endswith('.pdf'):
        messages.error(request, 'الملف المسموح به فقط: PDF')
        return redirect('training_coordinator_contractor_documents', contractor_user_id=contractor_user_id)

    content_type = (getattr(pdf_file, 'content_type', '') or '').lower()
    if content_type and 'pdf' not in content_type:
        messages.error(request, 'الملف المسموح به فقط: PDF')
        return redirect('training_coordinator_contractor_documents', contractor_user_id=contractor_user_id)

    ContractorDocument.objects.create(
        owner=contractor_user,
        title=title,
        pdf_file=pdf_file,
    )

    messages.success(request, 'تم رفع الملف بنجاح')
    return redirect('training_coordinator_contractor_documents', contractor_user_id=contractor_user_id)


@login_required
@require_POST
def training_coordinator_contractor_document_delete_action(request, contractor_user_id: int, doc_id: int):
    if not _training_coordinator_guard(request):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    coordinator = request.user
    contractor_user, _contractor_profile = _get_managed_contractor_or_404(coordinator, contractor_user_id)

    doc = get_object_or_404(ContractorDocument, id=doc_id, owner=contractor_user)
    linked_enrollments = list(doc.enrollments.all())
    for enrollment in linked_enrollments:
        try:
            enrollment.snapshot_supporting_documents([doc])
        except Exception:
            pass

    try:
        if doc.pdf_file:
            doc.pdf_file.delete(save=False)
    except Exception:
        pass

    doc.delete()
    messages.success(request, 'تم حذف الملف')
    return redirect('training_coordinator_contractor_documents', contractor_user_id=contractor_user_id)


def _trainer_reports_filtered_enrollments(request, user):
    """Return (qs, filter_context) for trainer reports without side effects."""
    base_qs = (
        EnrollmentRequest.objects
        .filter(Q(trainer__isnull=True) | Q(trainer=user))
        .select_related(
            'program',
            'contractor',
            'trainer',
            'exam_template',
            'contractor__contractorprofile',
        )
    )

    program_raw = (request.GET.get('program') or '').strip()
    status_raw = (request.GET.get('status') or '').strip()
    region_raw = (request.GET.get('region') or '').strip()
    q_raw = (request.GET.get('q') or '').strip()
    start_raw = (request.GET.get('start') or '').strip()
    end_raw = (request.GET.get('end') or '').strip()

    selected_program_id = None
    if program_raw:
        try:
            selected_program_id = int(program_raw)
        except (TypeError, ValueError):
            selected_program_id = None

    start_date = _parse_ymd_date(start_raw)
    end_date = _parse_ymd_date(end_raw)
    start_dt, end_dt = _date_range_to_aware_datetimes(start_date, end_date)

    qs = base_qs
    if selected_program_id:
        qs = qs.filter(program_id=selected_program_id)

    if status_raw:
        # Only accept valid statuses
        valid_statuses = {c[0] for c in EnrollmentRequest.Status.choices}
        if status_raw in valid_statuses:
            qs = qs.filter(status=status_raw)

    if region_raw:
        qs = qs.filter(contractor__contractorprofile__region=region_raw)

    if q_raw:
        qs = qs.filter(
            Q(contractor__username__icontains=q_raw)
            | Q(contractor__first_name__icontains=q_raw)
            | Q(contractor__last_name__icontains=q_raw)
            | Q(contractor__contractorprofile__company_name__icontains=q_raw)
            | Q(contractor__contractorprofile__id_number__icontains=q_raw)
        )

    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)

    # Annotate last exam attempt + certificate issuance
    last_attempt = (
        ExamAttempt.objects
        .filter(enrollment_id=OuterRef('pk'))
        .order_by('-completed_at', '-started_at', '-id')
    )
    cert_qs = Certificate.objects.filter(enrollment_id=OuterRef('pk'))

    qs = qs.annotate(
        last_score=Subquery(last_attempt.values('score')[:1]),
        last_passed=Subquery(last_attempt.values('passed')[:1]),
        last_attempt_completed_at=Subquery(last_attempt.values('completed_at')[:1]),
        has_certificate=Exists(cert_qs),
        certificate_issued_at=Subquery(cert_qs.values('issued_at')[:1]),
    )

    qs = qs.order_by('-created_at', '-id')

    filter_context = {
        'selected_program_id': selected_program_id,
        'selected_status': status_raw,
        'selected_region': region_raw,
        'q': q_raw,
        'start': start_raw,
        'end': end_raw,
    }
    return qs, filter_context


@login_required
def trainer_reports_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollments_qs, filter_ctx = _trainer_reports_filtered_enrollments(request, user)

    # Dropdowns
    programs = Program.objects.filter(is_active=True).order_by('title')
    regions = list(
        ContractorProfile.objects
        .exclude(region__isnull=True)
        .exclude(region__exact='')
        .values_list('region', flat=True)
        .distinct()
        .order_by('region')
    )
    region_label_map = {
        value: label
        for value, label in getattr(TrainingCoordinatorRegisterContractorForm, 'REGION_CHOICES', [])
        if value
    }
    regions_display = [
        {'value': region_value, 'label': region_label_map.get(region_value, region_value)}
        for region_value in regions
    ]
    statuses = list(EnrollmentRequest.Status.choices)

    total = enrollments_qs.count()
    status_counts = list(
        enrollments_qs.values('status').annotate(count=Count('id')).order_by('-count')
    )
    status_counts_map = {row['status']: row['count'] for row in status_counts}
    completed_count = (
        status_counts_map.get(EnrollmentRequest.Status.CERTIFIED, 0)
        + status_counts_map.get(EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD, 0)
    )

    # Certificates & SCORM completions in timeframe (reflect active filters)
    start_date = _parse_ymd_date(filter_ctx.get('start'))
    end_date = _parse_ymd_date(filter_ctx.get('end'))
    start_dt, end_dt = _date_range_to_aware_datetimes(start_date, end_date)

    enrollment_ids = enrollments_qs.values_list('id', flat=True)
    contractor_ids = enrollments_qs.values_list('contractor_id', flat=True).distinct()

    certs = Certificate.objects.filter(enrollment_id__in=enrollment_ids)
    scorm_certs = ScormCertificate.objects.filter(owner_id__in=contractor_ids)
    if start_dt:
        certs = certs.filter(issued_at__gte=start_dt)
        scorm_certs = scorm_certs.filter(issued_at__gte=start_dt)
    if end_dt:
        certs = certs.filter(issued_at__lte=end_dt)
        scorm_certs = scorm_certs.filter(issued_at__lte=end_dt)

    pass_cards_count = certs.filter(certificate_type=Certificate.CertificateType.PASS_CARD).count()
    issued_certificates_count = certs.filter(certificate_type=Certificate.CertificateType.CERTIFICATE).count()
    issued_scorm_certificates_count = scorm_certs.count()
    scorm_by_course = list(
        scorm_certs.values('course_name').annotate(count=Count('id')).order_by('-count', 'course_name')[:10]
    )

    return render(
        request,
        'accounts-templates/trainer-reports.html',
        {
            'programs': programs,
            'regions': regions_display,
            'statuses': statuses,
            'total': total,
            'pass_cards_count': pass_cards_count,
            'completed_count': completed_count,
            'status_counts': status_counts,
            'issued_certificates_count': issued_certificates_count,
            'issued_scorm_certificates_count': issued_scorm_certificates_count,
            'scorm_by_course': scorm_by_course,
            **filter_ctx,
        }
    )


@login_required
def trainer_registered_trainees_view(request):
    """Registered trainees page (filters + stats + list)."""
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollments_qs, filter_ctx = _trainer_reports_filtered_enrollments(request, user)

    programs = Program.objects.filter(is_active=True).order_by('title')
    regions = list(
        ContractorProfile.objects
        .exclude(region__isnull=True)
        .exclude(region__exact='')
        .values_list('region', flat=True)
        .distinct()
        .order_by('region')
    )
    region_label_map = {
        value: label
        for value, label in getattr(TrainingCoordinatorRegisterContractorForm, 'REGION_CHOICES', [])
        if value
    }
    regions_display = [
        {'value': region_value, 'label': region_label_map.get(region_value, region_value)}
        for region_value in regions
    ]
    statuses = list(EnrollmentRequest.Status.choices)

    total = enrollments_qs.count()
    status_counts = list(
        enrollments_qs.values('status').annotate(count=Count('id')).order_by('-count')
    )
    status_counts_map = {row['status']: row['count'] for row in status_counts}
    completed_count = (
        status_counts_map.get(EnrollmentRequest.Status.CERTIFIED, 0)
        + status_counts_map.get(EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD, 0)
    )

    start_date = _parse_ymd_date(filter_ctx.get('start'))
    end_date = _parse_ymd_date(filter_ctx.get('end'))
    start_dt, end_dt = _date_range_to_aware_datetimes(start_date, end_date)

    enrollment_ids = enrollments_qs.values_list('id', flat=True)
    contractor_ids = enrollments_qs.values_list('contractor_id', flat=True).distinct()

    certs = Certificate.objects.filter(enrollment_id__in=enrollment_ids)
    scorm_certs = ScormCertificate.objects.filter(owner_id__in=contractor_ids)
    if start_dt:
        certs = certs.filter(issued_at__gte=start_dt)
        scorm_certs = scorm_certs.filter(issued_at__gte=start_dt)
    if end_dt:
        certs = certs.filter(issued_at__lte=end_dt)
        scorm_certs = scorm_certs.filter(issued_at__lte=end_dt)

    pass_cards_count = certs.filter(certificate_type=Certificate.CertificateType.PASS_CARD).count()
    issued_certificates_count = certs.filter(certificate_type=Certificate.CertificateType.CERTIFICATE).count()
    issued_scorm_certificates_count = scorm_certs.count()

    return render(
        request,
        'accounts-templates/trainer-registered-trainees.html',
        {
            'programs': programs,
            'regions': regions_display,
            'statuses': statuses,
            'enrollments': enrollments_qs,
            'total': total,
            'pass_cards_count': pass_cards_count,
            'completed_count': completed_count,
            'status_counts': status_counts,
            'issued_certificates_count': issued_certificates_count,
            'issued_scorm_certificates_count': issued_scorm_certificates_count,
            **filter_ctx,
        }
    )


@login_required
def trainer_reports_export_csv(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollments_qs, filter_ctx = _trainer_reports_filtered_enrollments(request, user)

    # CSV (UTF-8 BOM for Excel)
    import csv

    start_label = (filter_ctx.get('start') or '').strip() or 'all'
    end_label = (filter_ctx.get('end') or '').strip() or 'all'
    filename = f"trainer-contractors-report_{start_label}_to_{end_label}.csv"

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp.write('\ufeff')

    writer = csv.writer(resp)
    writer.writerow([
        'Enrollment ID',
        'Program (Arabic)',
        'Program (English)',
        'Contractor Username',
        'Contractor Name',
        'Company',
        'Business Line',
        'ID Number',
        'Sadad Number',
        'Region',
        'Phone',
        'Registered At',
        'Status',
        'Status Display',
        'Exam Date',
        'Last Score',
        'Last Passed',
        'Certificate Issued',
        'Certificate Issued At',
    ])

    for e in enrollments_qs.iterator():
        contractor = getattr(e, 'contractor', None)
        profile = getattr(contractor, 'contractorprofile', None) if contractor else None
        program = getattr(e, 'program', None)

        contractor_name = ''
        if contractor:
            try:
                contractor_name = contractor.get_full_name() or contractor.username
            except Exception:
                contractor_name = getattr(contractor, 'username', '') or ''

        program_title_ar = (
            getattr(program, 'title_ar', None)
            or getattr(program, 'title', None)
            or getattr(program, 'name', None)
            or ''
        )
        program_title_en = (
            getattr(program, 'title_en', None)
            or getattr(program, 'title', None)
            or getattr(program, 'name', None)
            or ''
        )

        writer.writerow([
            getattr(e, 'id', ''),
            program_title_ar,
            program_title_en,
            getattr(contractor, 'username', '') if contractor else '',
            contractor_name,
            getattr(profile, 'company_name', '') if profile else '',
            getattr(profile, 'sec_business_line', '') if profile else '',
            getattr(profile, 'id_number', '') if profile else '',
            getattr(e, 'invoice_number', '') or '',
            getattr(profile, 'region', '') if profile else '',
            getattr(profile, 'phone_number', '') if profile else '',
            getattr(e, 'created_at', '') or '',
            getattr(e, 'status', '') or '',
            e.get_status_display() if hasattr(e, 'get_status_display') else '',
            getattr(e, 'exam_date', '') or '',
            getattr(e, 'last_score', '') if getattr(e, 'last_score', None) is not None else '',
            getattr(e, 'last_passed', '') if getattr(e, 'last_passed', None) is not None else '',
            'YES' if getattr(e, 'has_certificate', False) else 'NO',
            getattr(e, 'certificate_issued_at', '') or '',
        ])

    return resp


# =========================
# Contractor – My Requests
# =========================

@login_required
def contractor_requests_list_view(request):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    requests = (
        EnrollmentRequest.objects
        .filter(contractor=user, request_type=EnrollmentRequest.RequestType.INITIAL)
        .select_related('program', 'trainer', 'exam_template')
        .order_by('-created_at')
    )

    requests_rows = []
    for req in requests:
        requests_rows.append({
            'id': req.id,
            'program_name': _safe_program_name(getattr(req, 'program', None)),
            'trainer': getattr(req, 'trainer', None),
            'created_at': getattr(req, 'created_at', None),
            'status': getattr(req, 'status', None),
            'status_display': req.get_status_display() if hasattr(req, 'get_status_display') else '',
            'obj': req,
        })

    return render(
        request,
        'accounts-templates/contractor-requests.html',
        {'requests': requests, 'requests_rows': requests_rows}
    )


@login_required
def contractor_renewals_list_view(request):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    checked_certificate_id = None
    checked_raw = (request.GET.get('check') or '').strip()
    if checked_raw:
        try:
            checked_certificate_id = int(checked_raw)
        except (TypeError, ValueError):
            checked_certificate_id = None

    terminal_statuses = {
        EnrollmentRequest.Status.REJECTED,
        EnrollmentRequest.Status.FAILED,
        EnrollmentRequest.Status.CERTIFIED,
        EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD,
    }

    active_renewals_by_program = {}
    active_renewals = (
        EnrollmentRequest.objects
        .filter(contractor=user, request_type=EnrollmentRequest.RequestType.RENEWAL)
        .exclude(status__in=terminal_statuses)
        .order_by('-created_at', '-id')
    )
    for req in active_renewals:
        if req.program_id not in active_renewals_by_program:
            active_renewals_by_program[req.program_id] = req

    certificates = list(
        Certificate.objects
        .filter(owner=user)
        .select_related('program', 'enrollment')
        .order_by('-issued_at', '-id')
    )
    now = timezone.now()
    rows = []
    for cert in certificates:
        try:
            cert.ensure_expires_at()
            if cert.expires_at:
                cert.save(update_fields=['expires_at'])
        except Exception:
            pass

        renewal_months = 6
        try:
            renewal_months = int(getattr(getattr(cert, 'program', None), 'renewal_window_months', 6) or 6)
        except Exception:
            renewal_months = 6

        active_for_program = active_renewals_by_program.get(cert.program_id)
        eligible = False
        status_text = 'اضغط "التحقق من الأهلية" للتحقق'
        detail_text = ''
        renewal_start = cert.renewal_window_starts_at

        if checked_certificate_id == cert.id:
            if not cert.expires_at:
                status_text = 'غير مؤهل'
                detail_text = 'تعذر تحديد تاريخ انتهاء الشهادة.'
            elif now >= cert.expires_at:
                status_text = 'غير مؤهل'
                detail_text = 'الشهادة منتهية، يلزم تسجيل جديد في البرنامج.'
            elif active_for_program is not None:
                status_text = 'غير مؤهل'
                detail_text = f'يوجد طلب تجديد قائم بالفعل بحالة: {active_for_program.get_status_display()}.'
            elif renewal_start and now >= renewal_start:
                status_text = 'مؤهل للتجديد'
                detail_text = (
                    f'يمكنك رفع طلب التجديد الآن. نافذة التجديد: آخر {renewal_months} شهر '
                    f'قبل تاريخ الانتهاء {cert.expires_at.date()}.'
                )
                eligible = True
            else:
                status_text = 'غير مؤهل'
                detail_text = (
                    f'التجديد متاح خلال آخر {renewal_months} شهر فقط قبل الانتهاء. '
                    f'متاح من {renewal_start.date() if renewal_start else "—"}.'
                )

        rows.append({
            'certificate': cert,
            'is_checked': checked_certificate_id == cert.id,
            'is_eligible': eligible,
            'status_text': status_text,
            'detail_text': detail_text,
        })

    renewal_requests = (
        EnrollmentRequest.objects
        .filter(contractor=user, request_type=EnrollmentRequest.RequestType.RENEWAL)
        .select_related('program', 'trainer', 'exam_template')
        .order_by('-created_at')
    )

    return render(
        request,
        'accounts-templates/contractor-renewals.html',
        {
            'certificate_rows': rows,
            'renewal_requests': renewal_requests,
            'checked_certificate_id': checked_certificate_id,
        },
    )


# =========================
# Trainer – Review Requests
# =========================

def _trainer_requests_list_view_impl(request, *, request_type, request_scope: str, reset_url_name: str):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    base_qs = (
        EnrollmentRequest.objects
        .filter(Q(trainer__isnull=True) | Q(trainer=user), request_type=request_type)
        .select_related(
            'program',
            'contractor',
            'trainer',
            'exam_template',
            'contractor__contractorprofile',
        )
        .order_by('-created_at')
    )

    contractor_id = (request.GET.get('contractor_id') or '').strip()
    selected_region = (request.GET.get('region') or '').strip()
    program_raw = (request.GET.get('program') or '').strip()
    status_raw = (request.GET.get('status') or '').strip()

    selected_program_id = None
    if program_raw:
        try:
            selected_program_id = int(program_raw)
        except (TypeError, ValueError):
            selected_program_id = None

    requests = base_qs
    if contractor_id:
        requests = requests.filter(
            contractor__contractorprofile__id_number__icontains=contractor_id
        )

    if selected_program_id:
        requests = requests.filter(program_id=selected_program_id)

    if selected_region:
        requests = requests.filter(contractor__contractorprofile__region=selected_region)

    selected_status = ''
    if status_raw and status_raw in EnrollmentRequest.Status.values:
        selected_status = status_raw
        requests = requests.filter(status=selected_status)

    # Filter dropdown options (limited to trainer-visible requests)
    program_ids = list(base_qs.values_list('program_id', flat=True).distinct())
    programs_for_filter = Program.objects.filter(id__in=program_ids).order_by('title')

    regions_for_filter = list(
        ContractorProfile.objects
        .filter(user__enrollments__in=base_qs)
        .exclude(region__isnull=True)
        .exclude(region__exact='')
        .values_list('region', flat=True)
        .distinct()
        .order_by('region')
    )

    region_label_map = {
        value: label
        for value, label in getattr(TrainingCoordinatorRegisterContractorForm, 'REGION_CHOICES', [])
        if value
    }
    regions_for_filter_display = [
        {'value': region_value, 'label': region_label_map.get(region_value, region_value)}
        for region_value in regions_for_filter
    ]

    statuses_in_scope = set(base_qs.values_list('status', flat=True).distinct())
    status_label_map = {value: label for value, label in EnrollmentRequest.Status.choices}
    statuses_for_filter = [
        {'value': value, 'label': status_label_map.get(value, value)}
        for value in EnrollmentRequest.Status.values
        if value in statuses_in_scope
    ]

    return render(
        request,
        'accounts-templates/trainer-review-requests.html',
        {
            'requests': requests,
            'programs_for_filter': programs_for_filter,
            'regions_for_filter': regions_for_filter_display,
            'statuses_for_filter': statuses_for_filter,
            'contractor_id': contractor_id,
            'selected_program_id': selected_program_id,
            'selected_region': selected_region,
            'selected_status': selected_status,
            'request_scope': request_scope,
            'reset_url_name': reset_url_name,
        }
    )


@login_required
def trainer_requests_list_view(request):
    return _trainer_requests_list_view_impl(
        request,
        request_type=EnrollmentRequest.RequestType.INITIAL,
        request_scope='initial',
        reset_url_name='trainer_requests',
    )


@login_required
def trainer_renewal_requests_list_view(request):
    return _trainer_requests_list_view_impl(
        request,
        request_type=EnrollmentRequest.RequestType.RENEWAL,
        request_scope='renewal',
        reset_url_name='trainer_renewal_requests',
    )


# =========================
# ✅ Trainer – Exam Templates (PAGES)
# =========================

@login_required
def trainer_exam_grading_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    base_qs = (
        EnrollmentRequest.objects
        .filter(Q(trainer__isnull=True) | Q(trainer=user))
        .select_related('program', 'contractor', 'trainer')
        .order_by('-created_at')
    )

    grade_requests = []
    for enrollment in base_qs:
        gate = _trainer_external_assessment_gate(enrollment)
        if not gate['can_enter_external_assessment']:
            continue

        grade_requests.append({
            'enrollment': enrollment,
            'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
            'status_display': enrollment.get_status_display(),
        })

    selected_enrollment = None
    selected_gate = None
    selected_external_assessments = []
    selected_parts_scores = []
    selected_assessment_pdf_url = ''

    selected_raw = (request.GET.get('request_id') or '').strip()
    if selected_raw:
        try:
            selected_id = int(selected_raw)
        except (TypeError, ValueError):
            selected_id = None

        if selected_id:
            selected_enrollment = get_object_or_404(base_qs, id=selected_id)
            selected_gate = _trainer_external_assessment_gate(selected_enrollment)
            selected_external_assessments = list(
                ExternalPartAssessment.objects
                .filter(enrollment=selected_enrollment)
                .order_by('-submitted_at', '-id')
            )
            if selected_external_assessments:
                for assessment in selected_external_assessments:
                    pdf = getattr(assessment, 'pdf_file', None)
                    if pdf:
                        try:
                            selected_assessment_pdf_url = pdf.url
                            break
                        except Exception:
                            continue

            part_label_map = dict(ProgramExamPartConfig.PartType.choices)
            part_order = {
                ProgramExamPartConfig.PartType.THEORETICAL: 1,
                ProgramExamPartConfig.PartType.PRACTICAL: 2,
                ProgramExamPartConfig.PartType.PROJECT: 3,
            }
            required_parts = sorted(
                list(selected_gate['required_parts'] or []),
                key=lambda p: part_order.get(p, 99)
            )

            latest_theoretical_attempt = (
                selected_enrollment.exam_attempts
                .filter(completed_at__isnull=False)
                .order_by('-completed_at', '-id')
                .first()
            )

            external_by_part = {a.part_type: a for a in selected_external_assessments}

            for part_type in required_parts:
                part_label = part_label_map.get(part_type, part_type)
                score_text = '—'
                pass_text = 'غير متاح'

                if part_type == ProgramExamPartConfig.PartType.THEORETICAL:
                    if latest_theoretical_attempt and latest_theoretical_attempt.score is not None:
                        raw_score = float(latest_theoretical_attempt.score)
                        score_text = f'{raw_score:.2f}'.rstrip('0').rstrip('.')
                        pass_text = 'ناجح' if latest_theoretical_attempt.passed else 'راسب'
                else:
                    ext = external_by_part.get(part_type)
                    if ext is not None:
                        score_text = str(int(ext.grade_percent))
                        pass_text = 'ناجح' if ext.passed else 'راسب'

                selected_parts_scores.append({
                    'part_type': part_type,
                    'part_label': part_label,
                    'score_text': score_text,
                    'pass_text': pass_text,
                })

    return render(
        request,
        'accounts-templates/trainer-exam-grading.html',
        {
            'grade_requests': grade_requests,
            'selected_enrollment': selected_enrollment,
            'selected_external_assessments': selected_external_assessments,
            'selected_parts_scores': selected_parts_scores,
            'selected_assessment_pdf_url': selected_assessment_pdf_url,
            'selected_required_exam_parts': sorted(list(selected_gate['required_parts'])) if selected_gate else [],
            'selected_can_enter_external_assessment': bool(selected_gate and selected_gate['can_enter_external_assessment']),
        }
    )


@login_required
def trainer_exam_templates_list_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    templates = ExamTemplate.objects.filter(created_by=user).order_by('-id')

    # The template expects `templates` (list of ExamTemplate instances).
    return render(
        request,
        'accounts-templates/trainer-exam-templates.html',
        {
            'templates': templates,
        },
    )


@login_required
@require_http_methods(["GET"])
def trainer_exam_excel_template_download_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
    except ModuleNotFoundError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة')
        return redirect('trainer_exam_templates')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'

    headers = [
        'question',
        'type',
        'option_1',
        'option_2',
        'option_3',
        'option_4',
        'correct',
        'points',
        'explanation',
    ]
    ws.append(headers)
    ws['B1'].comment = Comment(
        'Allowed values: MCQ_SINGLE, MCQ_MULTI, TRUE_FALSE, SHORT_ANSWER, ESSAY',
        'System'
    )

    ws.append([
        'ما عاصمة المملكة العربية السعودية؟',
        'MCQ_SINGLE',
        'الرياض',
        'جدة',
        'الدمام',
        '',
        '1',
        '1',
        'العاصمة هي الرياض.',
    ])
    ws.append([
        'اختر اللغات البرمجية',
        'MCQ_MULTI',
        'Python',
        'HTML',
        'JavaScript',
        'CSS',
        '1,3',
        '2',
        '',
    ])
    ws.append([
        'السعودية تقع في قارة آسيا',
        'TRUE_FALSE',
        'صح',
        'خطأ',
        '',
        '',
        'A',
        '1',
        '',
    ])
    ws.append([
        'اذكر فائدة واحدة من التدريب',
        'SHORT_ANSWER',
        '',
        '',
        '',
        '',
        '',
        '1',
        '',
    ])

    column_widths = {
        'A': 50,
        'B': 20,
        'C': 26,
        'D': 26,
        'E': 26,
        'F': 26,
        'G': 14,
        'H': 10,
        'I': 40,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="trainer_exam_questions_template.xlsx"'
    return response


@login_required
def trainer_exam_template_create_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    if request.method == 'POST':
        form = ExamTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.created_by = user
            template.total_questions = 0
            template.save()

            messages.success(request, 'تم إنشاء قالب الاختبار ✅ الآن أضف الأسئلة')
            return redirect('trainer_exam_template_editor', template_id=template.id)
    else:
        form = ExamTemplateForm()

    return render(
        request,
        'accounts-templates/trainer-exam-template-editor.html',
        {
            'mode': 'create',
            'template': None,
            'template_obj': None,
            'template_form': form,
            'manual_formset': ManualQuestionFormSet(),
            'question_formset': ManualQuestionFormSet(),
            'excel_form': ExcelQuestionsUploadForm(),
            'questions': [],
        }
    )


@login_required
@transaction.atomic
@require_http_methods(["GET", "POST"])
def trainer_exam_template_editor_view(request, template_id: int):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    template = get_object_or_404(
        ExamTemplate.objects.select_related('created_by'),
        id=template_id,
        created_by=user
    )

    questions_qs = ExamQuestion.objects.filter(template=template).order_by('order', 'id')

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()

        # 1) تحديث بيانات القالب
        if action == 'update_template':
            template_form = ExamTemplateForm(request.POST, instance=template)
            if template_form.is_valid():
                template_form.save()
                messages.success(request, 'تم تحديث بيانات القالب ✅')
                return redirect('trainer_exam_template_editor', template_id=template.id)

            return render(
                request,
                'accounts-templates/trainer-exam-template-editor.html',
                {
                    'mode': 'edit',
                    'template': template,
                    'template_obj': template,
                    'template_form': template_form,
                    'manual_formset': ManualQuestionFormSet(),
                    'question_formset': ManualQuestionFormSet(),
                    'excel_form': ExcelQuestionsUploadForm(),
                    'questions': list(questions_qs),
                }
            )

        # 2) إضافة أسئلة يدويًا (أسئلة نصية فقط كبداية)
        if action == 'add_manual_questions':
            template_form = ExamTemplateForm(instance=template)
            manual_formset = ManualQuestionFormSet(request.POST)
            excel_form = ExcelQuestionsUploadForm()

            if manual_formset.is_valid():
                def _map_builder_qtype(raw: str):
                    v = (raw or '').strip().lower()
                    if v in {'multiple_choice_single', 'mcq_single', 'single', 'mcq'}:
                        return ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE
                    if v in {'multiple_choice_multi', 'mcq_multi', 'multi'}:
                        return ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI
                    if v in {'true_false', 'true/false', 'tf'}:
                        return ExamQuestion.QuestionType.TRUE_FALSE
                    if v in {'essay'}:
                        return ExamQuestion.QuestionType.ESSAY
                    return ExamQuestion.QuestionType.SHORT_ANSWER

                def _parse_int(v, default=1, min_v=1, max_v=100):
                    try:
                        x = int(v)
                    except (TypeError, ValueError):
                        return default
                    if x < min_v:
                        return min_v
                    if x > max_v:
                        return max_v
                    return x

                def _parse_indices(s: str):
                    if not s:
                        return []
                    parts = [p.strip() for p in str(s).replace(';', ',').split(',') if p.strip()]
                    out = []
                    for p in parts:
                        if p.isdigit():
                            out.append(int(p))
                            continue
                        if len(p) == 1 and p.upper() in {'A', 'B', 'C', 'D', 'E', 'F'}:
                            out.append(ord(p.upper()) - ord('A') + 1)
                    return out

                added = 0
                warnings_count = 0
                current_max_order = (
                    ExamQuestion.objects
                    .filter(template=template)
                    .aggregate(m=Max('order'))
                    .get('m') or 0
                )

                for form, item in zip(manual_formset.forms, manual_formset.cleaned_data):
                    q_text = (item.get('question_text') or '').strip()
                    if not q_text:
                        continue

                    prefix = getattr(form, 'prefix', '')
                    raw_type = request.POST.get(f'{prefix}-question_type', '')
                    question_type = _map_builder_qtype(raw_type)
                    points = _parse_int(request.POST.get(f'{prefix}-points', 1), default=1)
                    explanation = (request.POST.get(f'{prefix}-explanation', '') or '').strip()

                    # Options (A-D)
                    opt_texts = []
                    for i in range(1, 5):
                        txt = (request.POST.get(f'{prefix}-option_{i}', '') or '').strip()
                        if txt:
                            opt_texts.append(txt)

                    current_max_order += 1
                    q_obj = ExamQuestion.objects.create(
                        template=template,
                        order=current_max_order,
                        question_text=q_text,
                        question_type=question_type,
                        points=points,
                        explanation=explanation,
                    )

                    is_choice = question_type in {
                        ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE,
                        ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI,
                        ExamQuestion.QuestionType.TRUE_FALSE,
                    }

                    if is_choice:
                        # TRUE/FALSE: create default options if not provided
                        if question_type == ExamQuestion.QuestionType.TRUE_FALSE:
                            if len(opt_texts) < 2:
                                opt_texts = ['صح', 'خطأ']

                            correct_tf = (request.POST.get(f'{prefix}-correct_true_false', '') or '').strip().lower()
                            correct_text = 'صح' if correct_tf == 'true' else 'خطأ'
                            if correct_tf not in {'true', 'false'}:
                                warnings_count += 1

                            created = []
                            for idx, txt in enumerate(opt_texts[:2], start=1):
                                created.append(
                                    ExamOption(
                                        question=q_obj,
                                        order=idx,
                                        option_text=txt,
                                        is_correct=(txt == correct_text),
                                    )
                                )
                            ExamOption.objects.bulk_create(created)

                        else:
                            # MCQ: need at least 2 options; otherwise downgrade to SHORT_ANSWER
                            if len(opt_texts) < 2:
                                q_obj.question_type = ExamQuestion.QuestionType.SHORT_ANSWER
                                q_obj.save(update_fields=['question_type'])
                                warnings_count += 1
                            else:
                                created = []
                                for idx, txt in enumerate(opt_texts, start=1):
                                    created.append(
                                        ExamOption(
                                            question=q_obj,
                                            order=idx,
                                            option_text=txt,
                                            is_correct=False,
                                        )
                                    )
                                ExamOption.objects.bulk_create(created)

                                # mark correct
                                if question_type == ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE:
                                    correct_idx = _parse_int(
                                        request.POST.get(f'{prefix}-correct_option', ''),
                                        default=0,
                                        min_v=0,
                                        max_v=4,
                                    )
                                    if correct_idx <= 0:
                                        correct_idx = 1
                                        warnings_count += 1
                                    ExamOption.objects.filter(question=q_obj, order=correct_idx).update(is_correct=True)
                                else:
                                    idxs = _parse_indices(request.POST.get(f'{prefix}-correct_options', ''))
                                    if not idxs:
                                        idxs = [1]
                                        warnings_count += 1
                                    ExamOption.objects.filter(question=q_obj, order__in=idxs).update(is_correct=True)

                    added += 1

                total = ExamQuestion.objects.filter(template=template).count()
                if getattr(template, "total_questions", None) != total:
                    template.total_questions = total
                    template.save(update_fields=['total_questions'])

                if warnings_count:
                    messages.warning(request, f'تم الحفظ مع {warnings_count} تنبيه (بعض الأسئلة الاختيارية كانت ناقصة خيارات/إجابة صحيحة)')

                messages.success(request, f'تمت إضافة {added} سؤال ✅' if added else 'ما تم إضافة أي سؤال')
                return redirect('trainer_exam_template_editor', template_id=template.id)

            return render(
                request,
                'accounts-templates/trainer-exam-template-editor.html',
                {
                    'mode': 'edit',
                    'template': template,
                    'template_obj': template,
                    'template_form': template_form,
                    'manual_formset': manual_formset,
                    'question_formset': manual_formset,
                    'excel_form': excel_form,
                    'questions': list(questions_qs),
                }
            )

        # 3) رفع Excel واستيراد الأسئلة
        if action in {'upload_excel_questions', 'upload_excel'}:
            template_form = ExamTemplateForm(instance=template)
            manual_formset = ManualQuestionFormSet()
            excel_form = ExcelQuestionsUploadForm(request.POST, request.FILES)

            if excel_form.is_valid():
                excel_file = excel_form.cleaned_data['excel_file']
                replace_existing = bool(excel_form.cleaned_data.get('replace_existing'))
                default_question_type = (
                    excel_form.cleaned_data.get('default_question_type')
                    or ExamQuestion.QuestionType.SHORT_ANSWER
                )
                default_points = excel_form.cleaned_data.get('default_points') or 1

                try:
                    from openpyxl import load_workbook
                except ModuleNotFoundError:
                    messages.error(request, 'مكتبة openpyxl غير مثبتة. ثبّتها بالأمر: pip install openpyxl')
                    return redirect('trainer_exam_template_editor', template_id=template.id)

                try:
                    wb = load_workbook(excel_file, data_only=True)
                    ws = wb.active
                except Exception:
                    messages.error(request, 'تعذر قراءة ملف Excel (تأكد أنه .xlsx صحيح)')
                    return redirect('trainer_exam_template_editor', template_id=template.id)

                rows = []
                raw_rows = list(ws.iter_rows(min_row=1, values_only=True))
                for row in raw_rows:
                    if not row:
                        continue
                    # نطبعها لقيم نصية (مع الحفاظ على None)
                    normalized = []
                    for c in row:
                        if c is None:
                            normalized.append('')
                        else:
                            normalized.append(str(c).strip())
                    # تجاهل الصفوف الفاضية
                    if not any(x for x in normalized):
                        continue
                    rows.append(normalized)

                if not rows:
                    messages.error(request, 'الملف لا يحتوي بيانات صالحة')
                    return redirect('trainer_exam_template_editor', template_id=template.id)

                # دعم Header اختياري
                header = [c.strip().lower() for c in (rows[0] or [])]
                has_header = False
                if header and header[0] in {'question', 'questions', 'السؤال', 'الأسئلة', 'سؤال'}:
                    has_header = True

                if has_header:
                    header_map = {name: idx for idx, name in enumerate(header) if name}

                    def _find_col(*names):
                        for n in names:
                            if n in header_map:
                                return header_map[n]
                        return None

                    q_col = _find_col('question', 'questions', 'السؤال', 'سؤال')
                    correct_col = _find_col('correct', 'correct_answer', 'answer', 'الإجابة', 'الاجابة', 'صحيح')
                    type_col = _find_col('type', 'question_type', 'نوع', 'نوع السؤال')
                    points_col = _find_col('points', 'point', 'درجة', 'الدرجة')
                    explanation_col = _find_col('explanation', 'hint', 'شرح', 'تفسير')

                    option_cols = []
                    for name, idx in header_map.items():
                        if name.startswith('option') or name.startswith('اختيار') or name.startswith('خيار'):
                            option_cols.append(idx)
                    option_cols = sorted(set(option_cols))

                    data_rows = rows[1:]
                else:
                    # بدون Header: A سؤال, B..E خيارات (إن وجدت), F Correct (اختياري)
                    q_col = 0
                    option_cols = [1, 2, 3, 4]
                    correct_col = 5
                    type_col = None
                    points_col = None
                    explanation_col = None
                    data_rows = rows

                if replace_existing:
                    ExamQuestion.objects.filter(template=template).delete()

                current_max_order = (
                    ExamQuestion.objects
                    .filter(template=template)
                    .aggregate(m=Max('order'))
                    .get('m') or 0
                )

                def _parse_correct_marker(marker: str):
                    if marker is None:
                        return []
                    m = str(marker).strip()
                    if not m:
                        return []
                    # دعم "A" أو "B" أو "1" أو "1,3" ...
                    parts = [p.strip() for p in m.replace(';', ',').split(',') if p.strip()]
                    indices = []
                    for p in parts:
                        if p.isdigit():
                            indices.append(int(p))
                            continue
                        if len(p) == 1 and p.upper() in {'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'}:
                            indices.append(ord(p.upper()) - ord('A') + 1)
                            continue
                        # fallback: text marker
                        indices.append(p)
                    return indices

                def _parse_points(raw_value):
                    try:
                        val = int(float(str(raw_value).strip()))
                    except (TypeError, ValueError):
                        val = int(default_points)
                    if val < 1:
                        return 1
                    if val > 100:
                        return 100
                    return val

                added = 0
                for row in data_rows:
                    if not row:
                        continue
                    try:
                        q_text = (row[q_col] if q_col is not None and q_col < len(row) else '').strip()
                    except Exception:
                        q_text = ''

                    if not q_text:
                        continue

                    # اجمع الخيارات من الأعمدة المخصصة
                    option_texts = []
                    for idx in option_cols or []:
                        if idx < 0 or idx >= len(row):
                            continue
                        txt = (row[idx] or '').strip()
                        if txt:
                            option_texts.append(txt)

                    # نوع السؤال (اختياري)
                    qt = None
                    if type_col is not None and type_col < len(row):
                        qt_raw = (row[type_col] or '').strip()
                        if qt_raw:
                            qt = qt_raw

                    points_value = default_points
                    if points_col is not None and points_col < len(row):
                        points_value = _parse_points(row[points_col])

                    explanation_value = ''
                    if explanation_col is not None and explanation_col < len(row):
                        explanation_value = (row[explanation_col] or '').strip()

                    # correct marker (اختياري)
                    correct_marker = ''
                    if correct_col is not None and correct_col < len(row):
                        correct_marker = (row[correct_col] or '').strip()

                    # حدد نوع السؤال بشكل آمن
                    normalized_qt = (qt or '').strip().upper()
                    if normalized_qt in {'MCQ_MULTI', 'MULTI', 'MULTIPLE', 'MULTIPLE_CHOICE_MULTI'}:
                        question_type = ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI
                    elif normalized_qt in {'TRUE_FALSE', 'TF', 'TRUE/FALSE', 'صح/خطأ', 'صح خطأ'}:
                        question_type = ExamQuestion.QuestionType.TRUE_FALSE
                    elif normalized_qt in {'SHORT_ANSWER', 'SHORT', 'TEXT', 'إجابة قصيرة', 'نصي'}:
                        question_type = ExamQuestion.QuestionType.SHORT_ANSWER
                    elif normalized_qt in {'ESSAY', 'LONG', 'مقال', 'مقالية'}:
                        question_type = ExamQuestion.QuestionType.ESSAY
                    else:
                        if not normalized_qt:
                            if len(option_texts) >= 2:
                                if default_question_type in {
                                    ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE,
                                    ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI,
                                    ExamQuestion.QuestionType.TRUE_FALSE,
                                }:
                                    question_type = default_question_type
                                else:
                                    question_type = ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE
                            else:
                                question_type = default_question_type
                        else:
                            # لو النوع غير معروف في الملف: رجّع لسلوك آمن
                            question_type = (
                                ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE
                                if len(option_texts) >= 2
                                else ExamQuestion.QuestionType.SHORT_ANSWER
                            )

                    current_max_order += 1
                    q_obj = ExamQuestion.objects.create(
                        template=template,
                        order=current_max_order,
                        question_text=q_text,
                        question_type=question_type,
                        points=points_value,
                        explanation=explanation_value,
                    )

                    # إنشاء خيارات لو السؤال اختياري
                    is_choice = question_type in {
                        ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE,
                        ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI,
                        ExamQuestion.QuestionType.TRUE_FALSE,
                    }

                    if is_choice:
                        if question_type == ExamQuestion.QuestionType.TRUE_FALSE and len(option_texts) < 2:
                            option_texts = ['صح', 'خطأ']

                        created_opts = []
                        for idx, txt in enumerate(option_texts, start=1):
                            created_opts.append(
                                ExamOption(
                                    question=q_obj,
                                    order=idx,
                                    option_text=txt,
                                    is_correct=False,
                                )
                            )
                        if created_opts:
                            ExamOption.objects.bulk_create(created_opts)

                        # حاول تحديد الإجابة الصحيحة
                        markers = _parse_correct_marker(correct_marker)
                        if markers:
                            opts_qs = list(ExamOption.objects.filter(question=q_obj).order_by('order', 'id'))
                            by_order = {i + 1: o for i, o in enumerate(opts_qs)}
                            by_text = {o.option_text.strip(): o for o in opts_qs if (o.option_text or '').strip()}

                            to_mark = []
                            for m in markers:
                                if isinstance(m, int) and m in by_order:
                                    to_mark.append(by_order[m].id)
                                elif isinstance(m, str) and m in by_text:
                                    to_mark.append(by_text[m].id)

                            if to_mark:
                                ExamOption.objects.filter(id__in=to_mark).update(is_correct=True)

                    added += 1

                total = ExamQuestion.objects.filter(template=template).count()
                if getattr(template, "total_questions", None) != total:
                    template.total_questions = total
                    template.save(update_fields=['total_questions'])

                messages.success(request, f'تم استيراد {added} سؤال من Excel ✅')
                return redirect('trainer_exam_template_editor', template_id=template.id)

            return render(
                request,
                'accounts-templates/trainer-exam-template-editor.html',
                {
                    'mode': 'edit',
                    'template': template,
                    'template_obj': template,
                    'template_form': template_form,
                    'manual_formset': manual_formset,
                    'question_formset': manual_formset,
                    'excel_form': excel_form,
                    'questions': list(questions_qs),
                }
            )

        messages.error(request, 'الإجراء غير معروف')
        return redirect('trainer_exam_template_editor', template_id=template.id)

    return render(
        request,
        'accounts-templates/trainer-exam-template-editor.html',
        {
            'mode': 'edit',
            'template': template,
            'template_obj': template,
            'template_form': ExamTemplateForm(instance=template),
            'manual_formset': ManualQuestionFormSet(),
            'question_formset': ManualQuestionFormSet(),
            'excel_form': ExcelQuestionsUploadForm(),
            'questions': list(questions_qs),
        }
    )


@login_required
@require_POST
@transaction.atomic
def trainer_exam_question_delete_view(request, template_id: int, question_id: int):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    template = get_object_or_404(ExamTemplate, id=template_id, created_by=user)

    q = get_object_or_404(ExamQuestion, id=question_id, template=template)
    q.delete()

    total = ExamQuestion.objects.filter(template=template).count()
    if getattr(template, "total_questions", None) != total:
        template.total_questions = total
        template.save(update_fields=['total_questions'])

    messages.success(request, 'تم حذف السؤال ✅')
    return redirect('trainer_exam_template_editor', template_id=template.id)


# ==========================================================
# ✅✅✅ NEW: Trainer – Edit Question + Save Options (الحل الأساسي)
# ==========================================================

@login_required
@transaction.atomic
@require_http_methods(["GET", "POST"])
def trainer_exam_question_edit_view(request, template_id: int, question_id: int):
    """
    ✅ المدرب يعدّل السؤال + يضيف/يعدّل الخيارات
    ✅ الخيارات تنحفظ فعليًا في ExamOption model
    """
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    template = get_object_or_404(ExamTemplate, id=template_id, created_by=user)
    question = get_object_or_404(ExamQuestion, id=question_id, template=template)

    # queryset مرتب
    opt_qs = ExamOption.objects.filter(question=question).order_by('order', 'id')

    if request.method == "POST":
        q_form = ExamQuestionForm(request.POST, instance=question)
        opt_formset = ExamOptionInlineFormSet(request.POST, instance=question, queryset=opt_qs)

        if q_form.is_valid() and opt_formset.is_valid():
            q_obj = q_form.save()

            is_choice = q_obj.question_type in {
                ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE,
                ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI,
                ExamQuestion.QuestionType.TRUE_FALSE,
            }

            if not is_choice:
                # لو السؤال غير اختياري: نحذف الخيارات
                ExamOption.objects.filter(question=q_obj).delete()
            else:
                # ✅ نحفظ فقط الخيارات اللي فيها نص فعلي
                opts = opt_formset.save(commit=False)

                # نحذف اللي انطلب حذفها
                for obj in opt_formset.deleted_objects:
                    obj.delete()

                # نحفظ الجديدة/المعدلة بشرط النص موجود
                for obj in opts:
                    if (obj.option_text or '').strip():
                        obj.question = q_obj
                        obj.save()

                # ✅ ترتيب الخيارات حسب ترتيب الفورم
                if _has_field(ExamOption, 'order'):
                    new_order = 1
                    for f in opt_formset.forms:
                        cd = getattr(f, "cleaned_data", {}) or {}
                        if cd.get("DELETE"):
                            continue
                        txt = (cd.get("option_text") or "").strip()
                        inst = getattr(f, "instance", None)
                        if inst and inst.pk and txt:
                            ExamOption.objects.filter(pk=inst.pk).update(order=new_order)
                            new_order += 1

            messages.success(request, 'تم حفظ السؤال والخيارات ✅')
            return redirect('trainer_exam_template_editor', template_id=template.id)

        messages.error(request, 'تأكد من الحقول — يوجد أخطاء في السؤال أو الخيارات')

    else:
        q_form = ExamQuestionForm(instance=question)

        # ✅ تهيئة تلقائية لصح/خطأ (اختياري) إذا ما فيه خيارات
        initial = []
        if question.question_type == ExamQuestion.QuestionType.TRUE_FALSE and opt_qs.count() == 0:
            initial = [
                {'option_text': 'صح', 'is_correct': True},
                {'option_text': 'خطأ', 'is_correct': False},
            ]

        opt_formset = ExamOptionInlineFormSet(instance=question, queryset=opt_qs, initial=initial)

    return render(
        request,
        'accounts-templates/trainer-question-editor.html',
        {
            'template_obj': template,
            'question_obj': question,
            'q_form': q_form,
            'opt_formset': opt_formset,
        }
    )


# =========================
# ✅ Request Details (Contractor / Trainer)
# =========================

@login_required
def enrollment_request_detail_view(request, request_id: int):
    user = request.user
    is_contractor = _is_contractor(user)
    is_trainer = _is_trainer(user)
    is_coordinator = _is_training_coordinator(user)

    if not (is_contractor or is_trainer or is_coordinator):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    qs = EnrollmentRequest.objects.select_related('program', 'contractor', 'trainer', 'exam_template')

    # Role precedence: trainer > coordinator > contractor
    if is_trainer:
        qs = qs.filter(Q(trainer__isnull=True) | Q(trainer=user))
    elif is_coordinator:
        qs = qs.filter(contractor__contractorprofile__registered_by=user)
    elif is_contractor:
        qs = qs.filter(contractor=user)

    enrollment = get_object_or_404(qs, id=request_id)

    # Compute required parts once for all roles (used for gating UI/actions).
    gate = _trainer_external_assessment_gate(enrollment)
    required_parts = gate['required_parts']
    requires_theoretical = gate['requires_theoretical']
    requires_external = gate['requires_external']
    external_only = gate['external_only']
    locked_by_attempt_limit = gate['locked_by_attempt_limit']

    if is_trainer:
        template_name = 'accounts-templates/trainer-request-detail.html'

        can_review = enrollment.status == EnrollmentRequest.Status.NEW_REQUEST

        can_confirm_payment = (
            enrollment.status == EnrollmentRequest.Status.PAYMENT_VERIFICATION
            and bool(getattr(enrollment.program, 'requires_payment', False))
            and (enrollment.trainer is None or enrollment.trainer == user)
        )

        is_technical_exam = False
        try:
            is_technical_exam = enrollment.program.program_type == Program.ProgramType.TECHNICAL_EXAM
        except Exception:
            is_technical_exam = False

        # External-only exams behave like Technical Examination (date-only scheduling + manual grading).
        external_only = bool(requires_external and not requires_theoretical)
        is_technical_exam_ui = bool(is_technical_exam or external_only)

        can_schedule_exam = (
            (requires_theoretical or is_technical_exam or external_only)
            and enrollment.status == EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING
            and (enrollment.trainer is None or enrollment.trainer == user)
        )

        can_enter_external_assessment = gate['can_enter_external_assessment'] and (
            enrollment.trainer is None or enrollment.trainer == user
        )

        exam_templates = ExamTemplate.objects.filter(created_by=user).order_by('-id')

        external_assessments = list(
            ExternalPartAssessment.objects.filter(enrollment=enrollment).order_by('-submitted_at', '-id')
        )

        return render(
            request,
            template_name,
            {
                'enrollment': enrollment,
                'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
                'is_trainer': True,
                'is_contractor': False,
                'can_review': can_review,
                'can_confirm_payment': can_confirm_payment,
                'can_schedule_exam': can_schedule_exam,
                'exam_templates': exam_templates,
                'can_enter_external_assessment': can_enter_external_assessment,
                'external_assessments': external_assessments,
                'required_exam_parts': sorted(list(required_parts)),
                'is_technical_exam': is_technical_exam_ui,
                'requires_theoretical': requires_theoretical,
                'requires_external': requires_external,
                'external_only': external_only,
                'locked_by_attempt_limit': locked_by_attempt_limit,
                'max_attempts_per_part': MAX_EXAM_ATTEMPTS_PER_PART,
            }
        )

    template_name = 'accounts-templates/request-detail.html'
    return render(
        request,
        template_name,
        {
            'enrollment': enrollment,
            'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
            'is_contractor': bool(is_contractor and not is_coordinator),
            'is_trainer': False,
            'is_coordinator': bool(is_coordinator),
            'required_exam_parts': sorted(list(required_parts)),
            'requires_theoretical': requires_theoretical,
            'requires_external': requires_external,
            'external_only': external_only,
            'locked_by_attempt_limit': locked_by_attempt_limit,
            'max_attempts_per_part': MAX_EXAM_ATTEMPTS_PER_PART,
        }
    )


# =========================
# Trainer – Actions (Approve / Reject)
# =========================

@login_required
@transaction.atomic
def trainer_review_request_action(request, request_id):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects.select_for_update(),
        id=request_id
    )

    if enrollment.trainer and enrollment.trainer != user:
        messages.error(request, 'لا يمكنك التعديل على هذا الطلب')
        return redirect('trainer_requests')

    if request.method != 'POST':
        messages.error(request, 'طلب غير صالح')
        return redirect('trainer_requests')

    action = request.POST.get('action', '').strip()

    if enrollment.trainer is None:
        enrollment.trainer = user

    if action == 'approve':
        sadad_number = (request.POST.get('sadad_number') or request.POST.get('invoice_number') or '').strip()
        sadad_number = sadad_number.translate(str.maketrans('٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹', '01234567890123456789'))
        if getattr(enrollment.program, 'requires_payment', False) and not sadad_number:
            messages.error(request, 'يجب إدخال رقم سداد قبل اعتماد الطلب')
            return redirect('trainer_requests')

        if sadad_number and not sadad_number.isdigit():
            messages.error(request, 'رقم سداد يجب أن يحتوي على أرقام فقط')
            return redirect('trainer_requests')

        if sadad_number:
            enrollment.invoice_number = sadad_number

        enrollment.approve_by_trainer()
        enrollment.save(update_fields=['status', 'trainer', 'invoice_number'])

        messages.success(request, f'تم اعتماد الطلب وانتقل إلى مرحلة: {enrollment.get_status_display()}')
        return redirect('trainer_requests')

    if action == 'reject':
        reason = request.POST.get('rejection_reason', '').strip()
        if not reason:
            messages.error(request, 'يجب كتابة سبب الرفض')
            return redirect('trainer_requests')

        enrollment.status = EnrollmentRequest.Status.REJECTED
        enrollment.rejection_reason = reason
        enrollment.save(update_fields=['status', 'rejection_reason', 'trainer'])

        messages.success(request, 'تم رفض الطلب')
        return redirect('trainer_requests')

    messages.error(request, 'الإجراء غير معروف')
    return redirect('trainer_requests')


# =========================
# Contractor – Confirm Payment
# =========================

@login_required
@require_POST
@transaction.atomic
def contractor_confirm_payment_action(request, request_id: int):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects
        .select_for_update()
        .select_related('program', 'contractor', 'trainer'),
        id=request_id,
        contractor=user
    )

    if enrollment.status == EnrollmentRequest.Status.REJECTED:
        messages.error(request, 'لا يمكن تأكيد السداد لطلب مرفوض')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    if enrollment.status != EnrollmentRequest.Status.INVOICE_ISSUED:
        messages.error(request, 'لا يمكن تأكيد السداد في هذه المرحلة')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    if not getattr(enrollment.program, 'requires_payment', False):
        messages.error(request, 'هذا البرنامج لا يتطلب سداد')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    enrollment.submit_payment_by_contractor()
    enrollment.save(update_fields=['status'])

    messages.success(request, f'تم تأكيد السداد ✅ الحالة الحالية: {enrollment.get_status_display()}')
    return redirect('contractor_request_detail', request_id=enrollment.id)


# =========================
# Trainer – Confirm Payment
# =========================

@login_required
@require_POST
@transaction.atomic
def trainer_confirm_payment_action(request, request_id: int):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects
        .select_for_update()
        .select_related('program', 'trainer'),
        id=request_id
    )

    if enrollment.trainer and enrollment.trainer != user:
        messages.error(request, 'لا يمكنك التعديل على هذا الطلب')
        return redirect('trainer_requests')

    if enrollment.trainer is None:
        enrollment.trainer = user

    if enrollment.status != EnrollmentRequest.Status.PAYMENT_VERIFICATION:
        messages.error(request, 'لا يمكن تأكيد السداد في هذه المرحلة')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    if not getattr(enrollment.program, 'requires_payment', False):
        messages.error(request, 'هذا البرنامج لا يتطلب سداد')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    enrollment.confirm_payment()
    enrollment.save(update_fields=['status', 'trainer'])

    messages.success(request, f'تم تأكيد السداد ✅ الحالة الحالية: {enrollment.get_status_display()}')
    return redirect('trainer_request_detail', request_id=enrollment.id)


# =========================
# Trainer – External Part Assessment (Practical/Project)
# =========================


@login_required
@require_POST
@transaction.atomic
def trainer_external_assessment_submit_action(request, request_id: int):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects.select_for_update().select_related('program', 'contractor', 'trainer'),
        id=request_id
    )
    grading_detail_url = f"{reverse('trainer_exam_grading_view')}?request_id={enrollment.id}"

    if enrollment.trainer and enrollment.trainer != user:
        messages.error(request, 'لا يمكنك التعديل على هذا الطلب')
        return redirect('trainer_requests')

    if enrollment.trainer is None:
        enrollment.trainer = user

    # For Technical / external-only exams, only allow entering the grade after the exam has been
    # scheduled and confirmed by the contractor (EXAM_CONFIRMED).
    try:
        is_technical_exam = enrollment.program.program_type == Program.ProgramType.TECHNICAL_EXAM
    except Exception:
        is_technical_exam = False

    external_only = False
    try:
        required_gate = set(enrollment.required_exam_parts() or [])
        requires_theoretical_gate = ProgramExamPartConfig.PartType.THEORETICAL in required_gate
        requires_external_gate = bool(
            (ProgramExamPartConfig.PartType.PRACTICAL in required_gate)
            or (ProgramExamPartConfig.PartType.PROJECT in required_gate)
        )
        external_only = bool(requires_external_gate and not requires_theoretical_gate)
    except Exception:
        external_only = False

    allowed_statuses = {
        EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING,
        EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE,
        EnrollmentRequest.Status.FAILED,
    }
    if is_technical_exam or external_only:
        allowed_statuses = {
            EnrollmentRequest.Status.EXAM_CONFIRMED,
            EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE,
            EnrollmentRequest.Status.FAILED,
        }

    if enrollment.status not in allowed_statuses:
        messages.error(request, 'لا يمكن إدخال التقييم في هذه المرحلة')
        return redirect(grading_detail_url)

    if _is_locked_after_max_part_attempts(enrollment):
        enrollment.status = EnrollmentRequest.Status.FAILED
        enrollment.save(update_fields=['status', 'trainer'])
        messages.error(request, 'تم استنفاد محاولات الاختبار لهذا الطلب ويجب إعادة التسجيل من البداية.')
        return redirect(grading_detail_url)

    required_parts = set(enrollment.required_exam_parts() or [])
    has_practical = ProgramExamPartConfig.PartType.PRACTICAL in required_parts
    has_project = ProgramExamPartConfig.PartType.PROJECT in required_parts

    submitted_entries = []
    multi_submit_keys_present = (
        'practical_grade_percent' in request.POST
        or 'project_grade_percent' in request.POST
    )

    if has_practical and has_project and multi_submit_keys_present:
        def _read_multi_grade(field_name: str):
            raw = (request.POST.get(field_name) or '').strip()
            if not raw:
                return None
            try:
                value = int(raw)
            except (TypeError, ValueError):
                return None
            if value < 0 or value > 100:
                return None
            return value

        practical_grade = _read_multi_grade('practical_grade_percent')
        project_grade = _read_multi_grade('project_grade_percent')

        shared_pdf_file = request.FILES.get('pdf_file')

        if practical_grade is not None:
            submitted_entries.append((
                ProgramExamPartConfig.PartType.PRACTICAL,
                practical_grade,
                shared_pdf_file,
            ))
        if project_grade is not None:
            submitted_entries.append((
                ProgramExamPartConfig.PartType.PROJECT,
                project_grade,
                shared_pdf_file,
            ))

        if not submitted_entries:
            messages.error(request, 'أدخل درجة صحيحة (0-100) لجزء واحد على الأقل')
            return redirect(grading_detail_url)
    else:
        from .forms import ExternalPartAssessmentForm
        form = ExternalPartAssessmentForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, 'تحقق من بيانات التقييم')
            return redirect(grading_detail_url)

        part_type = (form.cleaned_data.get('part_type') or '').strip()
        grade = int(form.cleaned_data.get('grade_percent') or 0)
        submitted_entries.append((part_type, grade, form.cleaned_data.get('pdf_file')))

    # If theoretical is required (Solar), enforce passing theoretical first.
    if ProgramExamPartConfig.PartType.THEORETICAL in required_parts and not enrollment.has_passed_theoretical():
        messages.error(request, 'لا يمكن إدخال تقييم عملي/مشروع قبل اجتياز الجزء النظري')
        return redirect(grading_detail_url)

    part_results = []
    update_fields_set = {'status', 'trainer'}

    for part_type, grade, pdf_file in submitted_entries:
        if part_type == ProgramExamPartConfig.PartType.THEORETICAL:
            messages.error(request, 'لا يمكن إدخال تقييم للجزء النظري من هنا')
            return redirect(grading_detail_url)

        if part_type not in required_parts:
            messages.error(request, 'هذا الجزء غير مُدرج ضمن إعدادات البرنامج')
            return redirect(grading_detail_url)

        current_part_attempts = _get_part_attempts_count(enrollment, part_type)
        if current_part_attempts >= MAX_EXAM_ATTEMPTS_PER_PART:
            enrollment.status = EnrollmentRequest.Status.FAILED
            enrollment.save(update_fields=['status', 'trainer'])
            messages.error(request, 'تم استنفاد الحد الأقصى لمحاولات هذا الجزء. يجب إعادة التسجيل من البداية.')
            return redirect(grading_detail_url)

        passing = _part_passing_threshold(enrollment.program, part_type, default=60)
        passed = grade >= passing

        assessment, created = ExternalPartAssessment.objects.get_or_create(
            enrollment=enrollment,
            part_type=part_type,
            defaults={
                'grade_percent': grade,
                'passed': passed,
                'submitted_by': user,
                'pdf_file': pdf_file,
            }
        )

        if not created:
            assessment.grade_percent = grade
            assessment.passed = passed
            assessment.submitted_by = user
            if pdf_file:
                assessment.pdf_file = pdf_file
            assessment.save()

        part_attempt_field, part_attempts = _increment_part_attempts_count(enrollment, part_type)
        if part_attempt_field:
            update_fields_set.add(part_attempt_field)

        part_results.append({
            'part_type': part_type,
            'passed': passed,
            'attempts': part_attempts,
        })

    base_update_fields = list(update_fields_set)

    # Update enrollment status + issue outcome if ready
    if enrollment.can_issue_outcome():
        if enrollment.program.outcome_type == Program.OutcomeType.CERTIFICATE:
            enrollment.status = EnrollmentRequest.Status.CERTIFIED
            cert_type = Certificate.CertificateType.CERTIFICATE
        else:
            enrollment.status = EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD
            cert_type = Certificate.CertificateType.PASS_CARD

        verification_code = uuid.uuid4().hex[:12].upper()
        cert, _c_created = Certificate.objects.get_or_create(
            enrollment=enrollment,
            defaults={
                'owner': enrollment.contractor,
                'program': enrollment.program,
                'certificate_type': cert_type,
                'verification_code': verification_code,
            }
        )
        if not cert.verification_code:
            cert.verification_code = verification_code
            cert.save(update_fields=['verification_code'])

        try:
            pdf_bytes = _generate_certificate_pdf_bytes(
                owner_username=_contractor_english_name(enrollment.contractor),
                program_name=_program_english_name(enrollment.program),
                verification_code=cert.verification_code,
                issued_at=getattr(cert, 'issued_at', None) or timezone.now(),
                certificate_kind=getattr(cert, 'certificate_type', None) or cert_type,
            )
            filename = f"certificate_{enrollment.id}_{cert.verification_code}.pdf"
            cert.pdf_file.save(filename, ContentFile(pdf_bytes), save=False)
            cert.save()
        except Exception:
            pass

        enrollment.save(update_fields=base_update_fields)
        messages.success(request, 'تم حفظ التقييم وتم اعتماد النتيجة ✅')
        return redirect(grading_detail_url)

    # Not ready yet: allow up to 3 attempts per external part.
    first_failed_part = next((r for r in part_results if not r['passed']), None)
    if first_failed_part is not None:
        if first_failed_part['attempts'] >= MAX_EXAM_ATTEMPTS_PER_PART:
            enrollment.status = EnrollmentRequest.Status.FAILED
            enrollment.save(update_fields=base_update_fields)
            messages.error(request, 'تم حفظ التقييم — استنفد المتدرب 3 محاولات لهذا الجزء ويجب إعادة التسجيل من البداية')
            return redirect(grading_detail_url)

        enrollment.status = EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE
        enrollment.save(update_fields=base_update_fields)
        messages.error(
            request,
            f"تم حفظ التقييم — يوجد جزء غير مجتاز. المحاولة رقم {first_failed_part['attempts']} من {MAX_EXAM_ATTEMPTS_PER_PART}.",
        )
        return redirect(grading_detail_url)

    enrollment.status = EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE
    enrollment.save(update_fields=base_update_fields)
    messages.success(request, 'تم حفظ التقييم ✅')
    return redirect(grading_detail_url)


# =========================
# Contractor – Book Exam Slot (Confirm Attendance)
# =========================


@login_required
def trainer_exam_sessions_view(request):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    allowed_program_type_codes = [
        Program.ProgramType.COURSE_ATTENDANCE,
        Program.ProgramType.TECHNICAL_EXAM,
        Program.ProgramType.SOLAR_POWER_EXAM,
    ]
    selected_program_type = (request.GET.get('program_type') or request.POST.get('program_type') or '').strip()
    if selected_program_type not in allowed_program_type_codes:
        selected_program_type = ''

    choice_map = dict(Program.ProgramType.choices)
    program_type_choices = [(code, choice_map.get(code, code)) for code in allowed_program_type_codes]

    programs = Program.objects.all().order_by('title', 'id')
    if selected_program_type:
        programs = programs.filter(program_type=selected_program_type)
    program_map = {p.id: p for p in programs}

    def _exam_sessions_redirect_url(program_id: int | None = None) -> str:
        query = []
        if selected_program_type:
            query.append(f"program_type={selected_program_type}")
        if program_id:
            query.append(f"program={program_id}")
        base_url = reverse('trainer_exam_sessions_view')
        if not query:
            return base_url
        return f"{base_url}?{'&'.join(query)}"

    selected_program_id = (request.POST.get('program_id') or request.GET.get('program') or '').strip()
    selected_program = None
    if selected_program_id.isdigit():
        selected_program = program_map.get(int(selected_program_id))

    if request.method == 'POST':
        if selected_program is None:
            messages.error(request, 'اختر البرنامج أولاً')
            return redirect(_exam_sessions_redirect_url())

        required_parts = _required_exam_parts_for_program(selected_program)
        requires_theoretical = ProgramExamPartConfig.PartType.THEORETICAL in required_parts
        requires_external = bool(
            (ProgramExamPartConfig.PartType.PRACTICAL in required_parts)
            or (ProgramExamPartConfig.PartType.PROJECT in required_parts)
        )
        is_technical_exam = selected_program.program_type == Program.ProgramType.TECHNICAL_EXAM
        external_only = bool(requires_external and not requires_theoretical)

        if (not requires_theoretical) and (not is_technical_exam) and (not external_only):
            messages.error(request, 'هذا البرنامج لا يحتوي على اختبار يمكن جدولته')
            return redirect(_exam_sessions_redirect_url(selected_program.id))

        raw_dt = (request.POST.get('exam_date') or '').strip()
        if not raw_dt:
            messages.error(request, 'حدد موعد الاختبار أولاً')
            return redirect(_exam_sessions_redirect_url(selected_program.id))

        try:
            dt = datetime.fromisoformat(raw_dt)
        except ValueError:
            messages.error(request, 'صيغة التاريخ غير صحيحة')
            return redirect(_exam_sessions_redirect_url(selected_program.id))

        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        if dt <= timezone.now():
            messages.error(request, 'يجب ان يكون موعد الاختبار في المستقبل')
            return redirect(_exam_sessions_redirect_url(selected_program.id))

        exam_template = None
        code_hash = ''
        raw_code = ''

        if requires_theoretical and not (is_technical_exam or external_only):
            template_id = (request.POST.get('exam_template_id') or '').strip()
            if not template_id:
                messages.error(request, 'اختر قالب الاختبار')
                return redirect(_exam_sessions_redirect_url(selected_program.id))
            try:
                exam_template = ExamTemplate.objects.get(id=template_id, created_by=user)
            except ExamTemplate.DoesNotExist:
                messages.error(request, 'قالب الاختبار غير صحيح أو ليس من قوالبك')
                return redirect(_exam_sessions_redirect_url(selected_program.id))

            if ExamQuestion.objects.filter(template=exam_template).count() == 0:
                messages.error(request, 'القالب المختار لا يحتوي أسئلة')
                return redirect(_exam_sessions_redirect_url(selected_program.id))

            raw_code = (request.POST.get('exam_code') or '').strip()
            if not raw_code:
                messages.error(request, 'لازم تولد كود دخول الاختبار')
                return redirect(_exam_sessions_redirect_url(selected_program.id))
            if not raw_code.isdigit() or len(raw_code) != 8:
                messages.error(request, 'كود الاختبار يجب أن يكون 8 أرقام')
                return redirect(_exam_sessions_redirect_url(selected_program.id))
            code_hash = make_password(raw_code)

        ExamSession.objects.create(
            program=selected_program,
            created_by=user,
            exam_date=dt,
            exam_template=exam_template,
            exam_code_hash=code_hash,
            exam_code_plain=raw_code,
            is_active=True,
        )

        if raw_code:
            messages.success(request, f'تم إنشاء موعد اختبار بنجاح ✅ | كود الدخول: {raw_code}')
        else:
            messages.success(request, 'تم إنشاء موعد اختبار بنجاح ✅')
        return redirect(_exam_sessions_redirect_url(selected_program.id))

    selected_required_parts = set()
    selected_requires_theoretical = False
    selected_is_technical = False
    if selected_program is not None:
        selected_required_parts = _required_exam_parts_for_program(selected_program)
        selected_requires_theoretical = ProgramExamPartConfig.PartType.THEORETICAL in selected_required_parts
        selected_is_technical = selected_program.program_type == Program.ProgramType.TECHNICAL_EXAM

    sessions = (
        ExamSession.objects
        .filter(created_by=user, is_active=True, exam_date__gt=timezone.now())
        .select_related('program', 'exam_template', 'created_by')
        .order_by('exam_date', 'id')
    )
    if selected_program is not None:
        sessions = sessions.filter(program=selected_program)

    exam_templates = ExamTemplate.objects.filter(created_by=user).order_by('-id')
    return render(
        request,
        'accounts-templates/trainer-exam-sessions.html',
        {
            'programs': programs,
            'program_type_choices': program_type_choices,
            'selected_program_type': selected_program_type,
            'selected_program': selected_program,
            'sessions': sessions,
            'exam_templates': exam_templates,
            'requires_theoretical': selected_requires_theoretical,
            'is_technical_exam': selected_is_technical,
            'required_exam_parts': sorted(list(selected_required_parts)),
        }
    )


@login_required
def contractor_exam_sessions_view(request, request_id: int):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects.select_related('program', 'contractor'),
        id=request_id,
        contractor=user,
    )

    if enrollment.status == EnrollmentRequest.Status.REJECTED:
        messages.error(request, 'لا يمكن الوصول لمواعيد اختبار طلب مرفوض')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    allowed_statuses = {
        EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING,
        EnrollmentRequest.Status.EXAM_SCHEDULED,
    }
    if enrollment.status not in allowed_statuses:
        messages.error(request, 'لا يمكن حجز موعد اختبار في هذه المرحلة')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    sessions = (
        ExamSession.objects
        .filter(program=enrollment.program, is_active=True, exam_date__gt=timezone.now())
        .select_related('exam_template', 'created_by')
        .order_by('exam_date', 'id')
    )

    return render(
        request,
        'accounts-templates/contractor-exam-sessions.html',
        {
            'enrollment': enrollment,
            'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
            'sessions': sessions,
        }
    )


@login_required
@require_POST
@transaction.atomic
def contractor_book_exam_action(request, request_id: int):
    user = request.user
    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects
        .select_for_update()
        .select_related('program', 'contractor'),
        id=request_id,
        contractor=user,
    )

    if enrollment.status == EnrollmentRequest.Status.REJECTED:
        messages.error(request, 'لا يمكن الحجز لطلب مرفوض')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    valid_statuses_for_booking = {
        EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING,
        EnrollmentRequest.Status.EXAM_SCHEDULED,
    }
    if enrollment.status not in valid_statuses_for_booking:
        messages.error(request, 'لا يمكن حجز موعد الاختبار في هذه المرحلة')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    requires_theoretical = True
    try:
        required = set(enrollment.required_exam_parts() or [])
        requires_theoretical = ProgramExamPartConfig.PartType.THEORETICAL in required
    except Exception:
        requires_theoretical = True

    session_id_raw = (request.POST.get('session_id') or '').strip()
    if session_id_raw:
        try:
            session_id = int(session_id_raw)
        except ValueError:
            messages.error(request, 'موعد الاختبار غير صحيح')
            return redirect('contractor_exam_sessions_view', request_id=enrollment.id)

        exam_session = get_object_or_404(
            ExamSession.objects.select_related('program', 'exam_template', 'created_by'),
            id=session_id,
            program=enrollment.program,
            is_active=True,
        )

        if exam_session.exam_date <= timezone.now():
            messages.error(request, 'لا يمكن حجز موعد اختبار منتهٍ')
            return redirect('contractor_exam_sessions_view', request_id=enrollment.id)

        if requires_theoretical and not exam_session.exam_template_id:
            messages.error(request, 'هذا الموعد غير مكتمل (قالب الاختبار غير موجود)')
            return redirect('contractor_exam_sessions_view', request_id=enrollment.id)

        if requires_theoretical and not (exam_session.exam_code_hash or '').strip():
            messages.error(request, 'هذا الموعد غير مكتمل (كود دخول الاختبار غير موجود)')
            return redirect('contractor_exam_sessions_view', request_id=enrollment.id)

        enrollment.exam_session = exam_session
        enrollment.exam_date = exam_session.exam_date
        enrollment.exam_template = exam_session.exam_template
        enrollment.exam_code_hash = exam_session.exam_code_hash or ''
        if enrollment.trainer is None:
            enrollment.trainer = exam_session.created_by
        enrollment.status = EnrollmentRequest.Status.EXAM_CONFIRMED
        enrollment.save(update_fields=['exam_session', 'exam_date', 'exam_template', 'exam_code_hash', 'trainer', 'status'])

        messages.success(request, f'تم حجز موعد الاختبار ✅ الحالة الحالية: {enrollment.get_status_display()}')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    # Backward-compatible path: request-bound scheduled exam.
    if not enrollment.exam_date or (requires_theoretical and not enrollment.exam_template_id):
        messages.error(request, 'لم يتم إنشاء موعد الاختبار بعد')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    enrollment.status = EnrollmentRequest.Status.EXAM_CONFIRMED
    enrollment.save(update_fields=['status'])

    messages.success(request, f'تم حجز موعد الاختبار ✅ الحالة الحالية: {enrollment.get_status_display()}')
    return redirect('contractor_request_detail', request_id=enrollment.id)


# =========================
# Trainer – Schedule Exam (Template Required)
# =========================

@login_required
@require_POST
@transaction.atomic
def trainer_schedule_exam_action(request, request_id: int):
    user = request.user
    if not _is_trainer(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects
        .select_for_update()
        .select_related('program', 'trainer'),
        id=request_id
    )

    is_technical_exam = False
    try:
        is_technical_exam = enrollment.program.program_type == Program.ProgramType.TECHNICAL_EXAM
    except Exception:
        is_technical_exam = False

    # Determine if this is an online theoretical exam or an external-only (manual) exam.
    external_only = False
    try:
        required = set(enrollment.required_exam_parts() or [])
        requires_theoretical = ProgramExamPartConfig.PartType.THEORETICAL in required
        requires_external = bool(
            (ProgramExamPartConfig.PartType.PRACTICAL in required)
            or (ProgramExamPartConfig.PartType.PROJECT in required)
        )
        external_only = bool(requires_external and not requires_theoretical)

        if (not requires_theoretical) and (not is_technical_exam) and (not external_only):
            messages.error(request, 'هذا البرنامج لا يحتوي على اختبار نظري داخل النظام')
            return redirect('trainer_request_detail', request_id=enrollment.id)
    except Exception:
        external_only = False

    if enrollment.trainer and enrollment.trainer != user:
        messages.error(request, 'لا يمكنك التعديل على هذا الطلب')
        return redirect('trainer_requests')

    if enrollment.trainer is None:
        enrollment.trainer = user

    if enrollment.status != EnrollmentRequest.Status.WAITING_EXAM_SCHEDULING:
        messages.error(request, 'لا يمكن تحديد موعد الاختبار في هذه المرحلة')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    raw_dt = (request.POST.get('exam_date') or '').strip()
    if not raw_dt:
        messages.error(request, 'حدد موعد الاختبار أولاً')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    try:
        dt = datetime.fromisoformat(raw_dt)
    except ValueError:
        messages.error(request, 'صيغة التاريخ غير صحيحة')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())

    if dt <= timezone.now():
        messages.error(request, 'يجب ان يكون موعد الاختبار في المستقبل')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    # Technical / external-only exams: date-only scheduling (no online template / no access code).
    if is_technical_exam or external_only:
        enrollment.exam_date = dt
        enrollment.exam_template = None
        enrollment.exam_code_hash = ''
        enrollment.status = EnrollmentRequest.Status.EXAM_SCHEDULED
        enrollment.save(update_fields=['status', 'trainer', 'exam_date', 'exam_template', 'exam_code_hash'])
        messages.success(request, 'تم إنشاء موعد الاختبار بنجاح ✅')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    template_id = (request.POST.get('exam_template_id') or '').strip()
    if not template_id:
        messages.error(request, 'لازم تختار قالب اختبار قبل إنشاء الموعد')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    try:
        exam_template = ExamTemplate.objects.get(id=template_id, created_by=user)
    except ExamTemplate.DoesNotExist:
        messages.error(request, 'قالب الاختبار غير صحيح أو ليس من قوالبك')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    if ExamQuestion.objects.filter(template=exam_template).count() == 0:
        messages.error(request, 'القالب المختار ما فيه أسئلة. أضف أسئلة أولاً')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    raw_code = (request.POST.get('exam_code') or '').strip()
    if not raw_code:
        messages.error(request, 'لازم تولّد كود دخول للاختبار قبل حفظ الموعد')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    if not raw_code.isdigit() or len(raw_code) != 8:
        messages.error(request, 'كود الاختبار لازم يكون 8 أرقام فقط')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    code_hash = make_password(raw_code)

    # Prefer the model helper if it transitions state, but fall back to setting
    # fields directly after we've validated all inputs in this view.
    enrollment.schedule_exam(dt, exam_template=exam_template, exam_code_hash=code_hash)

    if enrollment.status != EnrollmentRequest.Status.EXAM_SCHEDULED or not enrollment.exam_date:
        enrollment.exam_date = dt
        enrollment.exam_template = exam_template
        enrollment.exam_code_hash = code_hash
        enrollment.status = EnrollmentRequest.Status.EXAM_SCHEDULED

    if enrollment.status != EnrollmentRequest.Status.EXAM_SCHEDULED or not enrollment.exam_date:
        messages.error(request, 'تعذر تحديد الموعد (تحقق من حالة الطلب)')
        return redirect('trainer_request_detail', request_id=enrollment.id)

    enrollment.save(update_fields=['status', 'trainer', 'exam_date', 'exam_template', 'exam_code_hash'])

    messages.success(request, f'تم إنشاء موعد الاختبار بنجاح ✅ | كود الدخول: {raw_code}')
    return redirect('trainer_request_detail', request_id=enrollment.id)


# =========================
# ✅ Contractor – Exam Page (GET/POST + Timer + Nav + Save)
# =========================

@login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def contractor_exam_view(request, request_id: int):
    user = request.user

    if not _is_contractor(user):
        messages.error(request, 'غير مصرح لك بالوصول')
        return redirect('home')

    enrollment = get_object_or_404(
        EnrollmentRequest.objects
        .select_for_update()
        .select_related('program', 'contractor', 'trainer', 'exam_template'),
        id=request_id,
        contractor=user
    )

    # برامج عملي/خارجي: لا تعرض صفحة الاختبار النظري
    try:
        if enrollment.program.program_type == Program.ProgramType.TECHNICAL_EXAM:
            messages.info(request, 'هذا البرنامج يحتوي على اختبار عملي فقط ويتم تقييمه من قبل المدرب.')
            return redirect('contractor_request_detail', request_id=enrollment.id)
    except Exception:
        pass

    if enrollment.status == EnrollmentRequest.Status.REJECTED:
        messages.error(request, 'لا يمكن الدخول للاختبار لأن الطلب مرفوض')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    if not enrollment.exam_template_id or not enrollment.exam_date:
        messages.error(request, 'لم يتم تحديد موعد الاختبار أو قالب الاختبار بعد')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    allowed_statuses = {
        EnrollmentRequest.Status.EXAM_CONFIRMED,
        EnrollmentRequest.Status.IN_EXAM,
        EnrollmentRequest.Status.FAILED,
    }
    if enrollment.status not in allowed_statuses:
        messages.error(request, 'لا يمكن الدخول للاختبار في هذه المرحلة')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    # If the request is locked as FAILED after max attempts on any required part, block re-entry.
    if enrollment.status == EnrollmentRequest.Status.FAILED and _is_locked_after_max_part_attempts(enrollment):
        messages.error(request, 'لا يمكن إعادة الاختبار — لم يتم اجتياز الاختبار ويجب التسجيل من جديد')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    now = timezone.now()
    if enrollment.exam_date and enrollment.exam_date > now:
        messages.error(request, 'الاختبار لم يبدأ بعد — انتظر موعد الاختبار المحدد')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    # ✅ Mandatory: exam must have an access code
    if not (getattr(enrollment, 'exam_code_hash', '') or '').strip():
        messages.error(request, 'لا يمكن الدخول للاختبار لأن كود الدخول غير مُنشأ بعد. تواصل مع المدرب.')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    # ✅ Exam code gate (mandatory if code exists)
    if getattr(enrollment, 'exam_code_hash', ''):
        unlock_key = _get_exam_code_unlock_session_key(enrollment.id)
        unlocked = bool(request.session.get(unlock_key, False))

        if not unlocked:
            if request.method == 'POST' and (request.POST.get('unlock_exam') or '').strip() == '1':
                entered = (request.POST.get('exam_code') or '').strip()
                if entered and entered.isdigit() and len(entered) == 8 and check_password(entered, enrollment.exam_code_hash):
                    request.session[unlock_key] = True
                    request.session.modified = True
                    messages.success(request, 'تم قبول كود الاختبار ✅ يمكنك الدخول الآن')
                    return redirect('contractor_exam_view', request_id=enrollment.id)

                messages.error(request, 'كود الاختبار غير صحيح')
                return render(
                    request,
                    'accounts-templates/contractor-exam-unlock.html',
                    {
                        'enrollment': enrollment,
                        'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
                    }
                )

            # GET (or unrelated POST): show unlock form
            return render(
                request,
                'accounts-templates/contractor-exam-unlock.html',
                {
                    'enrollment': enrollment,
                    'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
                }
            )

    # ✅ جلب الأسئلة
    questions = list(
        ExamQuestion.objects
        .filter(template=enrollment.exam_template)
        .order_by('order', 'id')
    )
    total = len(questions)

    if total == 0:
        messages.error(request, 'قالب الاختبار لا يحتوي أسئلة')
        return redirect('contractor_request_detail', request_id=enrollment.id)

    # ✅ جلب الخيارات (الحل: تظهر للمقاول لأنها محفوظة في ExamOption)
    q_ids = [q.id for q in questions]
    options_map = _options_map_for_questions(q_ids)

    # Attempt مفتوح
    attempt = (
        ExamAttempt.objects
        .filter(enrollment=enrollment, completed_at__isnull=True)
        .order_by('-id')
        .first()
    )
    if attempt is None:
        attempt = ExamAttempt.objects.create(enrollment=enrollment)

    # حول الحالة إلى IN_EXAM أول مرة فقط
    if enrollment.status == EnrollmentRequest.Status.EXAM_CONFIRMED:
        enrollment.status = EnrollmentRequest.Status.IN_EXAM
        enrollment.save(update_fields=['status'])

    # Timer
    duration_minutes = getattr(enrollment.exam_template, "duration_minutes", 30) or 30
    started_at = getattr(attempt, "started_at", None) or getattr(attempt, "created_at", None) or timezone.now()
    ends_at = started_at + timedelta(minutes=duration_minutes)
    remaining_seconds = int(max(0, (ends_at - timezone.now()).total_seconds()))

    # Session Answers
    session_key = _get_exam_session_key(enrollment.id)
    answers = request.session.get(session_key, {})  # {"<qid>": "<value or list>"}

    # =========================
    # POST: حفظ إجابة + تنقل
    # =========================
    if request.method == "POST":
        go = (request.POST.get("go") or "").strip()
        current_q = _normalize_q_index(request.POST.get("current_q"), total)

        qid = str(request.POST.get("question_id") or "").strip()

        if qid:
            updated = False

            answers_multi = request.POST.getlist("answers_multi")
            if answers_multi:
                answers[qid] = [str(x) for x in answers_multi if str(x).strip()]
                updated = True
            elif "answer_text" in request.POST:
                # textarea always posts (even if empty) — treat as user intent
                answers[qid] = (request.POST.get("answer_text") or "").strip()
                updated = True
            elif "answer" in request.POST:
                # radio posts only when selected
                answers[qid] = (request.POST.get("answer") or "").strip()
                updated = True

            if updated:
                request.session[session_key] = answers
                request.session.modified = True
                # Keep navigation auto-save silent (no message spam)
                if go == "stay":
                    messages.success(request, "تم حفظ الإجابة ✅")
        else:
            messages.error(request, "لم يتم تحديد السؤال")

        if go == "next":
            target = min(total, current_q + 1)
        elif go == "prev":
            target = max(1, current_q - 1)
        else:
            target = current_q

        return redirect(f"{reverse('contractor_exam_view', kwargs={'request_id': enrollment.id})}?q={target}")

    # =========================
    # GET: عرض سؤال حسب q
    # =========================
    current_index = _normalize_q_index(request.GET.get("q"), total)  # 1-based
    current_question = questions[current_index - 1]
    current_question_options = options_map.get(current_question.id, [])
    current_saved = answers.get(str(current_question.id))

    # answered set
    answered_ids = set()
    for k, v in answers.items():
        if v is None:
            continue
        if isinstance(v, list) and len(v) == 0:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        try:
            answered_ids.add(int(k))
        except (TypeError, ValueError):
            pass

    # compatibility for older template usage
    questions_rows = [{'q': q, 'options': options_map.get(q.id, [])} for q in questions]

    return render(
        request,
        'accounts-templates/contractor-exam.html',
        {
            'enrollment': enrollment,
            'program_name': _safe_program_name(getattr(enrollment, 'program', None)),
            'template_obj': enrollment.exam_template,

            'questions': questions,
            'total_questions': total,
            'current_index': current_index,
            'current_question': current_question,
            'current_question_options': current_question_options,
            'current_saved': current_saved,
            'answered_ids': answered_ids,
            'remaining_seconds': remaining_seconds,

            'questions_rows': questions_rows,
            'now': now,
            'can_start': True,
        }
    )

def _compute_attempt_score(enrollment: EnrollmentRequest, answers: dict) -> dict:
    """
    يرجع:
    - earned_points
    - total_points
    - score_percent
    - passed
    """
    template = enrollment.exam_template
    questions = list(
        ExamQuestion.objects.filter(template=template).order_by("order", "id")
    )

    # اجلب كل الخيارات مرة وحدة
    q_ids = [q.id for q in questions]
    options = list(ExamOption.objects.filter(question_id__in=q_ids).order_by("order", "id"))
    options_by_q = {}
    for o in options:
        options_by_q.setdefault(o.question_id, []).append(o)

    total_points = 0
    earned_points = 0

    for q in questions:
        q_points = int(getattr(q, "points", 1) or 1)
        total_points += q_points

        saved = answers.get(str(q.id))
        qtype = q.question_type

        # أسئلة اختيارية فقط نحسبها تلقائيًا
        is_choice = qtype in {
            ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE,
            ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI,
            ExamQuestion.QuestionType.TRUE_FALSE,
        }
        if not is_choice:
            continue

        opts = options_by_q.get(q.id, [])
        correct_ids = {str(o.id) for o in opts if o.is_correct}

        if qtype in {ExamQuestion.QuestionType.MULTIPLE_CHOICE_SINGLE, ExamQuestion.QuestionType.TRUE_FALSE}:
            chosen = (str(saved).strip() if saved is not None else "")
            if chosen and chosen in correct_ids:
                earned_points += q_points

        elif qtype == ExamQuestion.QuestionType.MULTIPLE_CHOICE_MULTI:
            chosen_list = []
            if isinstance(saved, list):
                chosen_list = [str(x).strip() for x in saved if str(x).strip()]
            elif isinstance(saved, str) and saved.strip():
                # احتياط لو جت string
                chosen_list = [saved.strip()]

            chosen_set = set(chosen_list)

            # صح فقط إذا التطابق كامل (نفس مجموعة الإجابات الصحيحة)
            if chosen_set and chosen_set == correct_ids:
                earned_points += q_points

    score_percent = 0.0
    if total_points > 0:
        score_percent = (earned_points / total_points) * 100.0

    passing = _part_passing_threshold(
        enrollment.program,
        ProgramExamPartConfig.PartType.THEORETICAL,
        default=60,
    )
    passed = score_percent >= passing

    return {
        "earned_points": earned_points,
        "total_points": total_points,
        "score_percent": round(score_percent, 2),
        "passed": passed,
        "passing_percent": passing,
    }


def _generate_certificate_pdf_bytes(
    owner_username: str,
    program_name: str,
    verification_code: str,
    issued_at: datetime | None = None,
    certificate_kind: str | None = None,
) -> bytes:
    """Generate branded PDF for certificate/pass card."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.colors import HexColor
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.graphics.barcode.qr import QrCodeWidget
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics import renderPDF
        import os
        import calendar
        import math
    except ImportError as e:
        raise RuntimeError("ReportLab is not installed. Install it with: pip install reportlab") from e

    def _try_register_ttf(font_name: str, paths: list[str]) -> bool:
        for p in paths:
            try:
                if p and os.path.exists(p):
                    pdfmetrics.registerFont(TTFont(font_name, p))
                    return True
            except Exception:
                continue
        return False

    def _add_years(dt: datetime, years: int) -> datetime:
        try:
            return dt.replace(year=dt.year + years)
        except ValueError:
            if dt.month == 2 and dt.day == 29:
                return dt.replace(year=dt.year + years, day=28)
            month_days = calendar.monthrange(dt.year + years, dt.month)[1]
            return dt.replace(year=dt.year + years, day=min(dt.day, month_days))

    def _draw_polygon(points: list[tuple[float, float]], fill_color):
        path = c.beginPath()
        path.moveTo(points[0][0], points[0][1])
        for px, py in points[1:]:
            path.lineTo(px, py)
        path.close()
        c.setFillColor(fill_color)
        c.drawPath(path, stroke=0, fill=1)

    font_regular = 'Helvetica'
    font_bold = 'Helvetica-Bold'
    try:
        reg_ok = _try_register_ttf(
            'AppFont',
            [
                r'C:\\Windows\\Fonts\\tahoma.ttf',
                r'C:\\Windows\\Fonts\\arial.ttf',
                r'C:\\Windows\\Fonts\\segoeui.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
            ],
        )
        bold_ok = _try_register_ttf(
            'AppFontBold',
            [
                r'C:\\Windows\\Fonts\\tahomabd.ttf',
                r'C:\\Windows\\Fonts\\arialbd.ttf',
                r'C:\\Windows\\Fonts\\segoeuib.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
            ],
        )
        if reg_ok:
            font_regular = 'AppFont'
            # if bold not available, reuse regular
            font_bold = 'AppFontBold' if bold_ok else 'AppFont'
    except Exception:
        pass

    issued_at = issued_at or timezone.now()
    expires_at = _add_years(issued_at, 3)
    issued_str = issued_at.strftime('%Y-%m-%d')
    expires_str = expires_at.strftime('%Y-%m-%d')
    owner_display = _english_text_only(owner_username, fallback='Contractor')
    program_display = _english_text_only(program_name, fallback='Program')

    kind = (certificate_kind or '').upper()
    is_pass_card = kind == 'PASS_CARD'
    document_title = 'SECO Employee Pass Card' if is_pass_card else 'Certificate of Completion'
    number_label = 'Pass Card No.' if is_pass_card else 'Certificate No.'

    # Brand colors
    orange = HexColor('#ff8c00')
    blue = HexColor('#164f95')
    navy = HexColor('#12325e')
    dark = HexColor('#1f2937')
    muted = HexColor('#6b7280')
    light_bg = HexColor('#f7fafc')
    soft_gray = HexColor('#d1d5db')
    white = HexColor('#ffffff')

    pagesize = landscape(A4)
    width, height = pagesize
    margin = 10 * mm

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=pagesize)

    # Base background
    c.setFillColor(light_bg)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    # Subtle diagonal lines
    c.setStrokeColor(HexColor('#ecf0f3'))
    c.setLineWidth(0.8)
    d = -height
    while d <= width:
        c.line(d, 0, d + height, height)
        d += 12

    # Top-left layered corner
    _draw_polygon([(0, height), (85 * mm, height), (43 * mm, height - 43 * mm), (0, height - 43 * mm)], soft_gray)
    _draw_polygon([(0, height), (52 * mm, height), (28 * mm, height - 24 * mm), (0, height - 24 * mm)], orange)
    _draw_polygon([(0, height), (42 * mm, height), (20 * mm, height - 20 * mm), (0, height - 20 * mm)], blue)

    # Bottom-right layered corner
    _draw_polygon([(width, 0), (width - 80 * mm, 0), (width - 40 * mm, 40 * mm), (width, 40 * mm)], soft_gray)
    _draw_polygon([(width, 0), (width - 47 * mm, 0), (width - 24 * mm, 24 * mm), (width, 24 * mm)], orange)
    _draw_polygon([(width, 0), (width - 36 * mm, 0), (width - 17 * mm, 17 * mm), (width, 17 * mm)], blue)

    # Main content panel
    c.setFillColor(white)
    c.setStrokeColor(HexColor('#dbe3ec'))
    c.setLineWidth(1.5)
    c.roundRect(margin, margin, width - 2 * margin, height - 2 * margin, 10, stroke=1, fill=1)

    # SEC logo (if available)
    logo_path = None
    try:
        logo_path = os.path.join(getattr(settings, 'MEDIA_ROOT', ''), 'SEC_logo.png')
        if not (logo_path and os.path.exists(logo_path)):
            logo_path = None
    except Exception:
        logo_path = None

    logo_drawn = False
    logo_x = None
    logo_bottom_y = None
    if logo_path:
        try:
            img = ImageReader(logo_path)
            iw, ih = img.getSize()
            target_h = 18 * mm
            target_w = (iw / max(ih, 1)) * target_h
            x = width - margin - target_w - 8 * mm
            y = height - margin - target_h - 7 * mm
            c.drawImage(img, x, y, width=target_w, height=target_h, mask='auto')
            logo_drawn = True
            logo_x = x
            logo_bottom_y = y
        except Exception:
            pass

    if is_pass_card:
        # Pass card title
        c.setFillColor(dark)
        c.setFont(font_bold, 24)
        c.drawString(margin + 18 * mm, height - margin - 24 * mm, document_title)

        # Decorative rosette
        c.saveState()
        c.setStrokeColor(HexColor('#b9d4ec'))
        c.setLineWidth(0.6)
        cx = width - margin - 52 * mm
        cy = height / 2
        for i in range(0, 360, 12):
            angle = i * math.pi / 180.0
            c.line(cx, cy, cx + 42 * mm * math.cos(angle), cy + 42 * mm * math.sin(angle))
        c.restoreState()

        # Details block
        x_left = margin + 18 * mm
        y0 = height - margin - 44 * mm
        row_gap = 12 * mm
        c.setFillColor(navy)
        c.setFont(font_bold, 14)
        c.drawString(x_left, y0, 'EMPLOYEE DETAILS')

        c.setFillColor(dark)
        c.setFont(font_regular, 13)
        c.drawString(x_left, y0 - row_gap, f'Name: {owner_display}')
        c.drawString(x_left, y0 - 2 * row_gap, f'Program: {program_display}')
        c.drawString(x_left, y0 - 3 * row_gap, f'{number_label}: {verification_code}')
        c.drawString(x_left, y0 - 4 * row_gap, f'Issue Date: {issued_str}')
        c.drawString(x_left, y0 - 5 * row_gap, f'Expiry Date: {expires_str}')
    else:
        # Certificate title
        content_left = margin + 20 * mm
        content_right = width - margin - 20 * mm
        if logo_drawn and logo_x is not None:
            # Keep title/subtitle away from top-right logo zone.
            content_right = min(content_right, logo_x - 8 * mm)
        title_center_x = (content_left + content_right) / 2
        title_y = height - margin - 34 * mm
        if logo_drawn and logo_bottom_y is not None:
            title_y = min(title_y, logo_bottom_y - 8 * mm)

        c.setFillColor(blue)
        c.setFont(font_bold, 36)
        c.drawCentredString(title_center_x, title_y, document_title)

        c.setFillColor(orange)
        c.setFont(font_bold, 14)
        c.drawCentredString(title_center_x, title_y - 13 * mm, 'HUMAN RESOURCES DEVELOPMENT HEREBY ACKNOWLEDGES')

        c.setFillColor(navy)
        c.setFont(font_bold, 26)
        c.drawString(margin + 22 * mm, height - margin - 74 * mm, f'NAME: {owner_display}')
        c.drawString(margin + 22 * mm, height - margin - 88 * mm, f'{number_label}: {verification_code}')

        c.setStrokeColor(soft_gray)
        c.setLineWidth(2)
        c.line(margin + 20 * mm, height - margin - 92 * mm, width - margin - 20 * mm, height - margin - 92 * mm)

        c.setFillColor(orange)
        c.setFont(font_bold, 18)
        c.drawString(margin + 22 * mm, height - margin - 104 * mm, 'IN RECOGNITION OF SATISFACTORY COMPLETION OF')

        c.setFillColor(dark)
        c.setFont(font_bold, 20)
        c.drawString(margin + 22 * mm, height - margin - 118 * mm, program_display)

        c.setFillColor(navy)
        c.setFont(font_bold, 14)
        c.drawString(margin + 22 * mm, margin + 22 * mm, f'Issue Date: {issued_str}')
        c.drawString(margin + 87 * mm, margin + 22 * mm, f'Expiry Date: {expires_str}')

    # QR code
    try:
        qr = QrCodeWidget(verification_code)
        bounds = qr.getBounds()
        qr_size = 24 * mm
        w = bounds[2] - bounds[0]
        h = bounds[3] - bounds[1]
        d = Drawing(qr_size, qr_size, transform=[qr_size / w, 0, 0, qr_size / h, 0, 0])
        d.add(qr)
        renderPDF.draw(d, c, width - margin - qr_size - 10 * mm, margin + 10 * mm)
        c.setFont(font_regular, 9)
        c.setFillColor(muted)
        c.drawRightString(width - margin - 10 * mm, margin + 8 * mm, 'Scan to verify')
    except Exception:
        pass

    # Signature line
    c.setStrokeColor(HexColor('#cbd5e1'))
    c.setLineWidth(1.2)
    sig_y = margin + 15 * mm
    c.line(margin + 20 * mm, sig_y, margin + 86 * mm, sig_y)
    c.setFillColor(muted)
    c.setFont(font_regular, 10)
    c.drawString(margin + 20 * mm, sig_y - 11, 'Authorized Signature')

    c.showPage()
    c.save()
    return buffer.getvalue()


@login_required
@require_POST
@transaction.atomic
def contractor_exam_submit_action(request, request_id: int):
    user = request.user
    if not _is_contractor(user):
        return JsonResponse({"ok": False, "message": "غير مصرح"}, status=403)

    enrollment = get_object_or_404(
        EnrollmentRequest.objects.select_for_update().select_related("program", "exam_template"),
        id=request_id,
        contractor=user
    )

    # ✅ enforce exam code unlock
    if getattr(enrollment, 'exam_code_hash', ''):
        unlock_key = _get_exam_code_unlock_session_key(enrollment.id)
        if not bool(request.session.get(unlock_key, False)):
            return JsonResponse({"ok": False, "message": "أدخل كود الاختبار أولاً"}, status=403)

    if not enrollment.exam_template_id:
        return JsonResponse({"ok": False, "message": "لا يوجد قالب اختبار"}, status=400)

    # ✅ اجلب محاولة مفتوحة
    attempt = (
        ExamAttempt.objects
        .filter(enrollment=enrollment, completed_at__isnull=True)
        .order_by("-id")
        .first()
    )
    if attempt is None:
        attempt = ExamAttempt.objects.create(enrollment=enrollment)

    session_key = _get_exam_session_key(enrollment.id)
    answers = request.session.get(session_key, {}) or {}

    # ✅ Save the current (last) question answer if included in finish request.
    # This prevents losing the final answer when the user ends the exam without navigating.
    qid = str(request.POST.get("question_id") or "").strip()
    if qid:
        updated = False
        answers_multi = request.POST.getlist("answers_multi")
        if answers_multi:
            answers[qid] = [str(x) for x in answers_multi if str(x).strip()]
            updated = True
        elif "answer_text" in request.POST:
            answers[qid] = (request.POST.get("answer_text") or "").strip()
            updated = True
        else:
            answer_single = (request.POST.get("answer") or "").strip()
            if answer_single:
                answers[qid] = answer_single
                updated = True

        if updated:
            request.session[session_key] = answers
            request.session.modified = True

    result = _compute_attempt_score(enrollment, answers)

    # ✅ اقفل المحاولة
    attempt.completed_at = timezone.now()
    attempt.score = float(result["score_percent"])
    attempt.passed = bool(result["passed"])
    attempt.save(update_fields=["completed_at", "score", "passed"])

    # ✅ عداد محاولات النظري (لكل جزء بشكل مستقل)
    next_attempts = int(getattr(enrollment, "attempts_count", 0) or 0) + 1
    enrollment.attempts_count = next_attempts  # backward-compatible legacy counter
    theoretical_attempt_field, theoretical_attempts = _increment_part_attempts_count(
        enrollment,
        ProgramExamPartConfig.PartType.THEORETICAL,
    )

    certificate_url = None
    certificate_error = None
    if result["passed"]:
        # ✅ إذا البرنامج يتطلب جزء عملي/خارجي، لا نصدر الشهادة الآن
        if not enrollment.can_issue_outcome():
            enrollment.status = EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE
        else:
            # ✅ تحديث الحالة حسب مخرج البرنامج
            if enrollment.program.outcome_type == Program.OutcomeType.CERTIFICATE:
                enrollment.status = EnrollmentRequest.Status.CERTIFIED
                cert_type = Certificate.CertificateType.CERTIFICATE
            else:
                enrollment.status = EnrollmentRequest.Status.COMPLETED_WITH_PASS_CARD
                cert_type = Certificate.CertificateType.PASS_CARD

            verification_code = uuid.uuid4().hex[:12].upper()

            cert, _created = Certificate.objects.get_or_create(
                enrollment=enrollment,
                defaults={
                    "owner": user,
                    "program": enrollment.program,
                    "certificate_type": cert_type,
                    "verification_code": verification_code,
                }
            )

            # لو موجود مسبقًا بس بدون verification code
            if not cert.verification_code:
                cert.verification_code = verification_code

            try:
                pdf_bytes = _generate_certificate_pdf_bytes(
                    owner_username=_contractor_english_name(user),
                    program_name=_program_english_name(enrollment.program),
                    verification_code=cert.verification_code,
                    issued_at=getattr(cert, 'issued_at', None) or timezone.now(),
                    certificate_kind=getattr(cert, 'certificate_type', None) or cert_type,
                )

                filename = f"certificate_{enrollment.id}_{cert.verification_code}.pdf"
                cert.pdf_file.save(filename, ContentFile(pdf_bytes), save=False)
                cert.save()

                if cert.pdf_file:
                    certificate_url = cert.pdf_file.url
            except Exception:
                certificate_url = None
                certificate_error = "تم اجتياز الاختبار، لكن تعذر إنشاء ملف PDF الآن. حاول لاحقًا أو تواصل مع الدعم."

    else:
        # Allow up to 3 attempts for the theoretical part, then lock as FAILED.
        if theoretical_attempts >= MAX_EXAM_ATTEMPTS_PER_PART:
            enrollment.status = EnrollmentRequest.Status.FAILED
        else:
            # Back to ready-for-exam stage for theoretical retake only.
            enrollment.status = EnrollmentRequest.Status.EXAM_CONFIRMED

    update_fields = ["status", "attempts_count"]
    if theoretical_attempt_field:
        update_fields.append(theoretical_attempt_field)
    enrollment.save(update_fields=update_fields)

    # ✅ صفّر إجابات السيشن عشان إعادة المحاولة تبدأ نظيفة
    request.session[session_key] = {}
    request.session.modified = True

    return JsonResponse({
        "ok": True,
        "passed": result["passed"],
        "score_percent": result["score_percent"],
        "earned_points": result["earned_points"],
        "total_points": result["total_points"],
        "passing_percent": result["passing_percent"],
        "certificate_url": certificate_url,
        "certificate_error": certificate_error,
        "waiting_external": (enrollment.status == EnrollmentRequest.Status.WAITING_PRACTICAL_GRADE),
        "attempts_count": enrollment.attempts_count,
        "part_attempts_count": theoretical_attempts,
        "max_attempts_per_part": MAX_EXAM_ATTEMPTS_PER_PART,
    })
